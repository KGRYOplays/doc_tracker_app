import os
import json
import base64
import sqlite3
import threading
import queue
import time
import atexit
import signal
from datetime import datetime
import pandas as pd
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC

# Files
DB_FILE = 'app.db'

# Periodic pull timer — reset to 0 so first pull fires immediately on startup
_periodic_last_run = 0
MASTER_DB = 'master_db.xlsx'
CODE_LOOKUP_EXCEL = 'code_lookup.xlsx'
ROUTING_RECORD_EXCEL = 'routing_record.xlsx'
OLD_LOG_EXCEL = 'scan_log.xlsx'

# Background Queue for Google Sheets Sync — replaced by persisted sync_queue table
gs_sync_queue = queue.Queue()  # kept only for backward-compat wakeup signal

# Last periodic pull info (for frontend notifications)
_last_pull_info = {"time": None, "status": "", "message": ""}
_last_pull_lock = threading.Lock()

# Event to wake the async worker when new items are enqueued
_sync_needed = threading.Event()

def update_last_pull_info(status, message):
    global _last_pull_info
    with _last_pull_lock:
        _last_pull_info = {"time": datetime.now().isoformat(), "status": status, "message": message}

def get_last_pull_info():
    with _last_pull_lock:
        return dict(_last_pull_info)

# ─────────────────────────────────────────────
#  CREDENTIAL ENCRYPTION HELPERS
# ─────────────────────────────────────────────

def _get_encryption_key():
    """Derive a Fernet encryption key from the app's secret key."""
    secret_key = os.environ.get('SECRET_KEY', 'barcode-routing-secret-2024-change-in-prod')
    kdf = PBKDF2HMAC(
        algorithm=hashes.SHA256(),
        length=32,
        salt=b'barcode_app_gsheets_salt_2024',
        iterations=480000,
    )
    key = base64.urlsafe_b64encode(kdf.derive(secret_key.encode()))
    return key


def encrypt_credentials(plaintext_json):
    """Encrypt Google service account credentials JSON for storage."""
    if not plaintext_json:
        return ''
    f = Fernet(_get_encryption_key())
    return f.encrypt(plaintext_json.encode()).decode()


def decrypt_credentials(encrypted_value):
    """Decrypt previously encrypted credentials. Returns empty string on failure."""
    if not encrypted_value:
        return ''
    try:
        f = Fernet(_get_encryption_key())
        return f.decrypt(encrypted_value.encode()).decode()
    except Exception:
        # If decryption fails, assume it's already plaintext (legacy data)
        return encrypted_value


def get_decrypted_setting(key, default=None):
    """Get a setting value, automatically decrypting 'gsheets_credentials'.
    
    Special handling for GSHEETS_CREDENTIALS env var: it is expected to be
    plaintext JSON (not encrypted), since the encryption key depends on
    SECRET_KEY which may differ between environments."""
    value = get_setting(key, default)
    if key == 'gsheets_credentials' and value:
        # Check if value is already valid JSON (env var path — plaintext)
        try:
            json.loads(value)
            return value  # Already plaintext JSON — return directly
        except (json.JSONDecodeError, TypeError):
            pass
        # Try decrypting (stored encrypted path)
        decrypted = decrypt_credentials(value)
        try:
            json.loads(decrypted)
            return decrypted
        except (json.JSONDecodeError, TypeError):
            pass
    return value


def save_encrypted_setting(key, value):
    """Save a setting, automatically encrypting 'gsheets_credentials'."""
    if key == 'gsheets_credentials' and value:
        # Validate it's valid JSON before encrypting
        try:
            parsed = json.loads(value)
            required_fields = ['type', 'project_id', 'private_key', 'client_email']
            missing = [f for f in required_fields if f not in parsed]
            if missing:
                raise ValueError(f"Missing required fields in credentials JSON: {', '.join(missing)}")
        except json.JSONDecodeError:
            raise ValueError("Credentials must be valid JSON")
        encrypted = encrypt_credentials(value)
        save_setting(key, encrypted)
    else:
        save_setting(key, value)


def get_db_connection():
    conn = sqlite3.connect(DB_FILE, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=30000")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn

def init_db():
    conn = get_db_connection()
    conn.execute("PRAGMA journal_mode=WAL")
    cursor = conn.cursor()
    
    # 1. Code Lookup Table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS code_lookup (
            code TEXT PRIMARY KEY,
            department TEXT NOT NULL,
            school_office TEXT NOT NULL,
            employees TEXT NOT NULL,
            doc_type TEXT,
            generated_at TEXT NOT NULL
        )
    ''')
    
    # 2. Routing Records Table — matches GSheet column order exactly
    # Dept, School/Office, Employee, Code, Doc Type, Remarks,
    # Receiving Office 1, Receiver Name 1, Timestamp 1, ... (x10)
    columns = [
        "id INTEGER PRIMARY KEY AUTOINCREMENT",
        "department TEXT NOT NULL",
        "school_office TEXT NOT NULL",
        "employee TEXT NOT NULL",
        "code TEXT NOT NULL",
        "doc_type TEXT DEFAULT ''",
        "remarks TEXT",
        "status TEXT NOT NULL DEFAULT 'for signature'",
    ]
    for i in range(1, 11):
        columns.append(f"receiving_office_{i} TEXT")
        columns.append(f"receiver_name_{i} TEXT")
        columns.append(f"timestamp_{i} TEXT")
    
    cursor.execute(f"CREATE TABLE IF NOT EXISTS routing_records ({', '.join(columns)})")
    cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_code_employee ON routing_records(code, employee)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_routing_school ON routing_records(school_office)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_routing_code ON routing_records(code)")
    
    # 3. Settings Table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    ''')
    
    # 4. Users Table (with school_office, email, requires_password_change columns)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'Pending',
            school_office TEXT DEFAULT '',
            email TEXT DEFAULT '',
            requires_password_change INTEGER DEFAULT 0,
            supervised_schools TEXT DEFAULT ''
        )
    ''')
    
    # Add columns if they don't exist (migration)
    try:
        cursor.execute("ALTER TABLE users ADD COLUMN school_office TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE users ADD COLUMN email TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE users ADD COLUMN requires_password_change INTEGER DEFAULT 0")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE users ADD COLUMN supervised_schools TEXT DEFAULT ''")
    except:
        pass
    try:
        cursor.execute("ALTER TABLE users ADD COLUMN employee_name TEXT DEFAULT ''")
    except:
        pass
    
    # 5. Master Data Table (replaces master_db.xlsx dependency)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS master_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department TEXT NOT NULL,
            school_office TEXT NOT NULL,
            employee_name TEXT NOT NULL
        )
    ''')
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_master_dept ON master_data(department)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_master_school ON master_data(school_office)")
    cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_master_unique ON master_data(department, school_office, employee_name)")
    
    # 6. Schools Table (RDBMS migration)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS schools (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            department TEXT NOT NULL DEFAULT ''
        )
    ''')
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_schools_name ON schools(name)")
    
    # 7. Employees Table (RDBMS migration)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            school_id INTEGER NOT NULL,
            FOREIGN KEY (school_id) REFERENCES schools(id)
        )
    ''')
    cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_employees_name_school ON employees(name, school_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_employees_school ON employees(school_id)")
    
    # 8. Code Employees Table (RDBMS migration — employee links for each code)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS code_employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT NOT NULL,
            employee_id INTEGER NOT NULL,
            FOREIGN KEY (code) REFERENCES code_lookup(code),
            FOREIGN KEY (employee_id) REFERENCES employees(id)
        )
    ''')
    cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_code_emp ON code_employees(code, employee_id)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_code_emp_code ON code_employees(code)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_code_emp_employee ON code_employees(employee_id)")
    
    # 9. Add employee_id column to routing_records (dual-write migration)
    try:
        cursor.execute("ALTER TABLE routing_records ADD COLUMN employee_id INTEGER REFERENCES employees(id)")
    except Exception:
        pass
    
    # 10. Pending profile changes (admin approval table)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS pending_profile_changes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            changes_json TEXT NOT NULL,
            requested_at TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'pending'
        )
    ''')
    
    conn.commit()
    
    # Seed default admin user if users table is empty
    import hashlib, secrets
    cursor.execute("SELECT COUNT(*) FROM users")
    if cursor.fetchone()[0] == 0:
        admin_password = secrets.token_urlsafe(12)
        admin_hash = hashlib.sha256(admin_password.encode()).hexdigest()
        cursor.execute("INSERT INTO users (username, password_hash, role, status, school_office, email, requires_password_change) VALUES (?, ?, ?, ?, ?, ?, ?)",
                       ('admin', admin_hash, 'Admin', 'Approved', '', 'admin@deped.gov.ph', 1))
        conn.commit()
        print(f"!!! DEFAULT ADMIN PASSWORD: {admin_password} - CHANGE IMMEDIATELY !!!")
    
    # Seed default sync token if not present
    cursor.execute("SELECT value FROM settings WHERE key = 'sync_token'")
    if not cursor.fetchone():
        import secrets
        default_token = secrets.token_urlsafe(32)
        cursor.execute("INSERT INTO settings (key, value) VALUES (?, ?)", ('sync_token', default_token))
        conn.commit()
    
    # Seed GSheets settings from environment variables if missing from SQLite.
    # This is critical on Render: after a wipe, SQLite is blank but env vars persist.
    # We persist them into SQLite so the rest of the app can use get_setting() normally.
    _seed_gsheets_settings_from_env(cursor, conn)
    
    # Seed master_data from Excel file if master_data table is empty and Excel exists
    cursor.execute("SELECT COUNT(*) FROM master_data")
    if cursor.fetchone()[0] == 0 and os.path.exists(MASTER_DB):
        try:
            print("Seeding master_data from all sheets in master_db.xlsx...")
            xl = pd.ExcelFile(MASTER_DB)
            
            # 1. Load Sheet1 data and build a mapping of school -> department
            school_dept_map = {}
            sheet1_seeded = 0
            if 'Sheet1' in xl.sheet_names:
                df1 = pd.read_excel(xl, 'Sheet1')
                if all(col in df1.columns for col in ['Department', 'School/Office', 'Employee_Name']):
                    for _, row in df1.iterrows():
                        dept = str(row.get('Department', '')).strip()
                        school = str(row.get('School/Office', '')).strip()
                        emp = str(row.get('Employee_Name', '')).strip()
                        if dept and school and emp:
                            cursor.execute(
                                "INSERT OR IGNORE INTO master_data (department, school_office, employee_name) VALUES (?, ?, ?)",
                                (dept, school, emp)
                            )
                            if cursor.rowcount > 0:
                                sheet1_seeded += 1
                            school_key = school.lower().strip()
                            if school_key not in school_dept_map:
                                school_dept_map[school_key] = dept
                    conn.commit()
                    print(f"Seeded {sheet1_seeded} entries from Sheet1.")
            
            # 2. Load Sheet3 data (Senior High Schools list)
            shs_schools = set()
            if 'Sheet3' in xl.sheet_names:
                df3 = pd.read_excel(xl, 'Sheet3')
                # Read all values in the sheet to find schools matching senior high
                for col in df3.columns:
                    for val in df3[col].dropna():
                        val_str = str(val).strip()
                        if val_str:
                            shs_schools.add(val_str.lower())
            
            # 3. Load Sheet2 data (Additional detailed employee list)
            sheet2_seeded = 0
            if 'Sheet2' in xl.sheet_names:
                df2 = pd.read_excel(xl, 'Sheet2')
                required_cols = ['FIRSTNAME', 'LASTNAME', 'SCHOOL/ OFFICE']
                if all(col in df2.columns for col in required_cols):
                    for _, row in df2.iterrows():
                        # Construct employee name
                        name_parts = [
                            row.get('FIRSTNAME'),
                            row.get('MIDDLE INITIAL'),
                            row.get('LASTNAME'),
                            row.get('SUFFIX')
                        ]
                        emp = " ".join([str(n).strip() for n in name_parts if pd.notna(n) and str(n).strip()])
                        school = str(row.get('SCHOOL/ OFFICE', '')).strip()
                        
                        if not emp or not school:
                            continue
                            
                        # Determine department
                        school_key = school.lower().strip()
                        dept = school_dept_map.get(school_key)
                        
                        if not dept:
                            if school_key in shs_schools or '(shs)' in school_key:
                                dept = 'SENIOR HIGH SCHOOL'
                            elif 'elementary' in school_key or ' es' in school_key:
                                dept = 'ELEMENTARY SCHOOL'
                            elif 'high school' in school_key or ' hs' in school_key or 'integrated' in school_key:
                                dept = 'SECONDARY SCHOOL'
                            else:
                                dept = 'SDO - MANILA'
                                
                        cursor.execute(
                            "INSERT OR IGNORE INTO master_data (department, school_office, employee_name) VALUES (?, ?, ?)",
                            (dept, school, emp)
                        )
                        if cursor.rowcount > 0:
                            sheet2_seeded += 1
                    conn.commit()
                    print(f"Seeded {sheet2_seeded} entries from Sheet2.")
                    
            cursor.execute("SELECT COUNT(*) FROM master_data")
            total_master = cursor.fetchone()[0]
            print(f"Total master_data loaded: {total_master} entries.")
        except Exception as e:
            print(f"Note: Could not seed from master_db.xlsx sheets: {e}")
    
    # 4. Migrate: add receiver_name columns to existing databases
    _migrate_add_receiver_names(conn)
    
    # 5. Migrate: add status column to existing routing_records tables
    try:
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(routing_records)")
        existing_cols = {row[1] for row in cursor.fetchall()}
        if 'status' not in existing_cols:
            cursor.execute("ALTER TABLE routing_records ADD COLUMN status TEXT NOT NULL DEFAULT 'for signature'")
            conn.commit()
            print("Migration: added 'status' column to routing_records.")
    except Exception as e:
        print(f"Migration note: could not add status column: {e}")
    
    # 6. Create persisted sync_queue table (replaces in-memory gs_sync_queue)
    try:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS sync_queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                table_name TEXT NOT NULL,
                ref_id INTEGER,
                code TEXT,
                employee TEXT,
                operation TEXT NOT NULL DEFAULT 'upsert',
                created_at TEXT NOT NULL,
                retry_count INTEGER DEFAULT 0,
                last_error TEXT,
                status TEXT DEFAULT 'pending'
            )
        """)
        conn.commit()
        print("Migration: sync_queue table ready.")
    except Exception as e:
        print(f"Migration note: could not create sync_queue table: {e}")
    
    conn.close()
    
    # Run data migrations (if old excel files exist and db is empty)
    migrate_from_excel()
    
    # No startup GSheets calls — persisted sync_queue replays on worker start.
    # With upsert-only pulls, local-only rows survive, so no pre-push is needed
    # and no full restore is required. The first user request serves from SQLite
    # immediately (zero API cost on startup).

    # Start background GSheets worker for async scan sync.
    # Scans queue row_ids; worker reads from SQLite and pushes to GSheets in background.
    start_gsheets_worker()

    # Start periodic (every N minutes) pull from GSheets so SQLite always has
    # reasonably fresh data even without a server restart. The interval defaults
    # to 5 minutes and can be changed at runtime via the gsheets_pull_interval setting.
    start_periodic_gsheets_pull()
    
    # Register graceful shutdown handler — tries to flush pending sync items
    # before exit. The persisted queue items survive even if this doesn't complete.
    atexit.register(_flush_sync_queue_on_shutdown)
    try:
        signal.signal(signal.SIGTERM, _signal_handler)
    except (ValueError, AttributeError):
        pass  # SIGTERM not available on some platforms (e.g., Windows)

def _flush_sync_queue_on_shutdown():
    """Best-effort flush of pending sync items on graceful shutdown.
    If this doesn't complete (e.g., SIGKILL), items persist in the
    sync_queue table and will be replayed on next startup."""
    print("Shutdown: flushing pending sync queue...")
    enabled = get_setting('gsheets_enabled') == 'True'
    sheet_id = get_setting('gsheets_id')
    creds = get_decrypted_setting('gsheets_credentials')
    if not enabled or not sheet_id or not creds:
        return
    try:
        conn = get_db_connection()
        pending = conn.execute("SELECT COUNT(*) FROM sync_queue WHERE status = 'pending'").fetchone()[0]
        if pending:
            print(f"Shutdown: processing {pending} pending sync items...")
            _process_pending_sync(sheet_id, creds)
        conn.close()
        print("Shutdown: sync queue flushed.")
    except Exception as e:
        print(f"Shutdown: could not flush sync queue: {e}")

def _signal_handler(signum, frame):
    """SIGTERM handler — triggers shutdown flush."""
    _flush_sync_queue_on_shutdown()

def _seed_gsheets_settings_from_env(cursor, conn):
    """Seed critical GSheets settings from environment variables into SQLite.
    
    Called during init_db() so that even after a Render wipe (which blanks SQLite),
    the app can immediately connect to Google Sheets using env vars set in the
    Render dashboard. Values are only written if the key does not already exist.
    """
    env_map = {
        'gsheets_enabled':     ('GSHEETS_ENABLED',     'True'),
        'gsheets_id':          ('GSHEETS_ID',           None),
        'gsheets_credentials': ('GSHEETS_CREDENTIALS',  None),
    }
    seeded = []
    for setting_key, (env_name, default_val) in env_map.items():
        # Only seed if not already in SQLite
        existing = cursor.execute(
            "SELECT value FROM settings WHERE key = ?", (setting_key,)
        ).fetchone()
        if not existing:
            env_val = os.environ.get(env_name, default_val)
            if env_val:
                cursor.execute(
                    "INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)",
                    (setting_key, env_val)
                )
                seeded.append(setting_key)
    if seeded:
        conn.commit()
        print(f"Seeded GSheets settings from env vars: {', '.join(seeded)}")


def _migrate_add_receiver_names(conn):
    """Add receiver_name_N columns if they don't exist (for upgrading old databases)."""
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(routing_records)")
    existing_cols = [row[1] for row in cursor.fetchall()]
    
    # Add remarks column if missing
    if 'remarks' not in existing_cols:
        try:
            cursor.execute("ALTER TABLE routing_records ADD COLUMN remarks TEXT")
            print("Migration: added remarks column to routing_records")
        except Exception as e:
            print(f"Migration warning (remarks): {e}")
    
    # Find all stage numbers from receiving_office_N columns
    stages = set()
    for col in existing_cols:
        if col.startswith('receiving_office_'):
            try:
                stages.add(int(col.split('_')[-1]))
            except:
                pass
    
    changed = False
    for stage in sorted(stages):
        col = f"receiver_name_{stage}"
        if col not in existing_cols:
            try:
                cursor.execute(f"ALTER TABLE routing_records ADD COLUMN {col} TEXT")
                changed = True
                print(f"Migration: added {col} to routing_records")
            except Exception as e:
                print(f"Migration warning (non-fatal): {e}")
    
    if changed:
        conn.commit()


def ensure_routing_columns(conn, stage):
    """Ensure routing_records has all 3 columns for the given stage."""
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(routing_records)")
    existing_cols = [row[1] for row in cursor.fetchall()]
    
    col_office   = f"receiving_office_{stage}"
    col_receiver = f"receiver_name_{stage}"
    col_ts       = f"timestamp_{stage}"
    
    changed = False
    for col in [col_office, col_receiver, col_ts]:
        if col not in existing_cols:
            try:
                cursor.execute(f"ALTER TABLE routing_records ADD COLUMN {col} TEXT")
                changed = True
            except Exception as e:
                print(f"Error adding column {col}: {e}")
    
    if changed:
        conn.commit()
        print(f"Dynamically expanded SQLite schema -> Stage {stage} columns added.")
    return changed


# Environment variable names that map to setting keys.
# These are checked as fallbacks when the SQLite settings table is empty
# (e.g. after a Render free-tier wipe). Set these in Render > Environment.
_ENV_FALLBACKS = {
    'gsheets_enabled':     'GSHEETS_ENABLED',      # e.g. "True"
    'gsheets_id':          'GSHEETS_ID',            # e.g. "1BxiM..."
    'gsheets_credentials': 'GSHEETS_CREDENTIALS',   # plaintext service-account JSON
}


def get_setting(key, default=None):
    """Read a setting from SQLite, falling back to environment variables for
    critical GSheets config when the database has been wiped by Render."""
    conn = get_db_connection()
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    conn.close()
    if row:
        return row['value']
    # Fallback: check environment variable
    env_name = _ENV_FALLBACKS.get(key)
    if env_name:
        env_val = os.environ.get(env_name)
        if env_val:
            return env_val
    return default


def save_setting(key, value):
    conn = get_db_connection()
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
    conn.commit()
    conn.close()


# ---- MASTER DATA HELPERS (replaces master_db.xlsx dependency) ----

def get_all_departments():
    """Return sorted list of unique departments from master_data."""
    conn = get_db_connection()
    rows = conn.execute("SELECT DISTINCT department FROM master_data ORDER BY department").fetchall()
    conn.close()
    return [r['department'] for r in rows]


def get_schools_for_department(department):
    """Return sorted list of unique schools for a given department."""
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT DISTINCT school_office FROM master_data WHERE department = ? ORDER BY school_office",
        (department,)
    ).fetchall()
    conn.close()
    return [r['school_office'] for r in rows]


def get_employees_for_school(department, school):
    """Return sorted list of employee names for a given department + school.
    Names are normalized and deduplicated to prevent format duplicates."""
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT DISTINCT employee_name FROM master_data WHERE department = ? AND school_office = ?",
        (department, school)
    ).fetchall()
    conn.close()
    seen = set()
    unique = []
    for r in rows:
        norm = reorder_name(r['employee_name'])
        norm_key = normalize_name_str(norm)
        if norm_key not in seen:
            seen.add(norm_key)
            unique.append(norm)
    unique.sort()
    return unique


def get_all_schools():
    """Return sorted list of all unique schools/offices across all departments."""
    conn = get_db_connection()
    rows = conn.execute("SELECT DISTINCT school_office FROM master_data ORDER BY school_office").fetchall()
    conn.close()
    return [r['school_office'] for r in rows]


def get_employees_by_school(school):
    """Return all employees belonging to a specific school (across all departments).
    Names are normalized via reorder_name() and deduplicated to prevent
    'LastName, FirstName' and 'FirstName LastName' from both appearing."""
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT DISTINCT employee_name FROM master_data WHERE school_office = ?",
        (school,)
    ).fetchall()
    conn.close()
    # Normalize and deduplicate: reorder names, then use a set
    seen = set()
    unique = []
    for r in rows:
        norm = reorder_name(r['employee_name'])
        norm_key = normalize_name_str(norm)
        if norm_key not in seen:
            seen.add(norm_key)
            unique.append(norm)
    unique.sort()
    return unique


def get_employees_with_suggestions(department, school):
    """Return a dict with employees split into same-school vs other-schools.
    
    Returns: {
        "same_school": ["Name1", "Name2", ...],
        "other_schools": [{"name": "Name3", "school": "School X"}, ...]
    }
    Used by the unified employee picker to provide suggestions in priority order.
    """
    same_school = get_employees_for_school(department, school)
    same_school_set = {normalize_name_str(e) for e in same_school}

    # Get all employees from master_data, excluding this school
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT DISTINCT employee_name, school_office FROM master_data WHERE school_office != ?",
        (school,)
    ).fetchall()
    conn.close()

    other_schools = []
    seen = set()
    for r in rows:
        norm = reorder_name(r['employee_name'])
        norm_key = normalize_name_str(norm)
        if norm_key not in seen and norm_key not in same_school_set:
            seen.add(norm_key)
            other_schools.append({"name": norm, "school": r['school_office']})

    other_schools.sort(key=lambda x: x['name'])
    return {"same_school": same_school, "other_schools": other_schools}


def add_master_entries(department, school, employees):
    """Add multiple employees under a department+school. employees is a list of names."""
    conn = get_db_connection()
    cursor = conn.cursor()
    count = 0
    for emp in employees:
        if emp.strip():
            cursor.execute(
                "INSERT OR IGNORE INTO master_data (department, school_office, employee_name) VALUES (?, ?, ?)",
                (department.strip(), school.strip(), emp.strip())
            )
            if cursor.rowcount > 0:
                count += 1
    conn.commit()
    conn.close()
    return count


def delete_master_entry(entry_id):
    """Delete a single master_data entry by id."""
    conn = get_db_connection()
    conn.execute("DELETE FROM master_data WHERE id = ?", (entry_id,))
    conn.commit()
    conn.close()


def get_all_master_data():
    """Return all master_data entries for admin management."""
    conn = get_db_connection()
    rows = conn.execute("SELECT * FROM master_data ORDER BY department, school_office, employee_name").fetchall()
    conn.close()
    return [dict(r) for r in rows]


# --- END MASTER DATA HELPERS ---


# --- EMPLOYEE NAME FORMAT HELPERS ---
import re

# Pattern: "AND 9 OTHERS", "AND 12 OTHERS", etc.
_AND_OTHERS_RE = re.compile(r'^(.*?)\s+AND\s+(\d+)\s+OTHERS$', re.IGNORECASE)


def normalize_name_str(name):
    """Normalize a name for comparison: lowercase, strip periods, collapse spaces."""
    s = name.strip().lower()
    s = s.replace('.', '')
    s = re.sub(r'\s+', ' ', s)
    return s


def reorder_name(name):
    """Convert 'LastName, FirstName MI' to 'FirstName MI LastName'.
    If no comma, return as-is.
    Also handles 'LastName, FirstName MI AND N OTHERS' → 'FirstName MI LastName AND N OTHERS'."""
    name = name.strip()
    if ',' not in name:
        return name

    parts = [p.strip() for p in name.split(',', 1)]
    if len(parts) != 2:
        return name

    last_name = parts[0].strip()
    rest = parts[1].strip()

    # Check for "AND N OTHERS" suffix
    m = _AND_OTHERS_RE.match(rest)
    if m:
        first_part = m.group(1).strip()
        num_others = m.group(2)
        return f"{first_part} {last_name} AND {num_others} OTHERS"

    # Simple "LastName, FirstName MI" → "FirstName MI LastName"
    return f"{rest} {last_name}"


def migrate_routing_employee_names():
    """Rewrite all routing_records.employee from 'LastName, FirstName MI' to 'FirstName MI LastName'.
    Returns (rewritten_count, total_count, samples) where samples is list of 'old → new' strings."""
    conn = get_db_connection()
    cursor = conn.cursor()
    rows = cursor.execute("SELECT id, employee FROM routing_records").fetchall()
    total = len(rows)
    rewritten = 0
    samples = []

    for row in rows:
        old_name = row['employee']
        new_name = reorder_name(old_name)
        if new_name != old_name:
            cursor.execute("UPDATE routing_records SET employee = ? WHERE id = ?", (new_name, row['id']))
            rewritten += 1
            if len(samples) < 5:
                samples.append(f"{old_name} → {new_name}")

    conn.commit()
    conn.close()
    print(f"Migrated {rewritten}/{total} routing employee names to new format.")
    return rewritten, total, samples


# --- REAL-TIME EXCEL EXPORT ---
def trigger_excel_export():
    """Export SQLite data to local Excel files via openpyxl (no pandas in hot path)."""
    def _export():
        try:
            from openpyxl import Workbook
            conn = sqlite3.connect(DB_FILE)
            conn.row_factory = sqlite3.Row

            # ── Code Lookup ──
            wb = Workbook()
            ws = wb.active
            ws.title = "Code Lookup"
            ws.append(["Code", "Department", "School/Office", "Employees", "Doc Type", "Generated"])
            cur = conn.execute(
                "SELECT code, department, school_office, employees, doc_type, generated_at "
                "FROM code_lookup ORDER BY generated_at DESC"
            )
            for row in cur:
                ws.append([row['code'], row['department'], row['school_office'],
                           row['employees'], row['doc_type'], row['generated_at']])
            wb.save(CODE_LOOKUP_EXCEL)

            # ── Routing Records ──
            cur.execute("PRAGMA table_info(routing_records)")
            cols = [r[1] for r in cur.fetchall()]
            stages = sorted(set(
                int(c.split("_")[-1]) for c in cols
                if c.startswith("receiving_office_")
            )) or list(range(1, 11))

            header = ["Department", "School/Office", "Employee", "Code", "Doc Type", "Remarks"]
            for i in stages:
                header.append(f"Receiving Office {i}")
                if f"receiver_name_{i}" in cols:
                    header.append(f"Receiver Name {i}")
                header.append(f"Timestamp {i}")

            wb2 = Workbook()
            ws2 = wb2.active
            ws2.title = "Routing Records"
            ws2.append(header)

            cur.execute("SELECT * FROM routing_records ORDER BY id DESC")
            for row in cur:
                vals = [row['department'], row['school_office'], row['employee'],
                        row['code'], row.get('doc_type', '') or '',
                        row.get('remarks', '') or '']
                for i in stages:
                    vals.append(row.get(f'receiving_office_{i}', '') or '')
                    if f"receiver_name_{i}" in cols:
                        vals.append(row.get(f'receiver_name_{i}', '') or '')
                    vals.append(row.get(f'timestamp_{i}', '') or '')
                ws2.append(vals)

            wb2.save(ROUTING_RECORD_EXCEL)
            conn.close()
            print("Excel databases successfully updated (openpyxl).")
        except Exception as e:
            print(f"Error exporting to Excel (openpyxl): {e}")

    threading.Thread(target=_export, daemon=True).start()


# --- DATA MIGRATION FROM OLD EXCEL ---
def migrate_from_excel():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. Migrate code lookup
    cursor.execute("SELECT COUNT(*) FROM code_lookup")
    if cursor.fetchone()[0] == 0 and os.path.exists(CODE_LOOKUP_EXCEL):
        try:
            df = pd.read_excel(CODE_LOOKUP_EXCEL)
            for _, row in df.iterrows():
                code     = str(row.get('Code', '')).strip()
                dept     = str(row.get('Department', 'Unknown')).strip()
                school   = str(row.get('School/Office', 'Unknown')).strip()
                employees = str(row.get('Employees', '')).strip()
                doc_type  = str(row.get('Doc Type', '')).strip()
                gen       = str(row.get('Generated', datetime.now().strftime("%m/%d/%Y"))).strip()
                if code:
                    cursor.execute(
                        "INSERT OR IGNORE INTO code_lookup (code, department, school_office, employees, doc_type, generated_at) "
                        "VALUES (?, ?, ?, ?, ?, ?)",
                        (code, dept, school, employees, doc_type, gen)
                    )
            conn.commit()
            print("Migrated code lookup database successfully.")
        except Exception as e:
            print(f"Error migrating code lookup: {e}")
    
    # 2. Migrate routing records
    cursor.execute("SELECT COUNT(*) FROM routing_records")
    if cursor.fetchone()[0] == 0:
        if os.path.exists(ROUTING_RECORD_EXCEL):
            try:
                df = pd.read_excel(ROUTING_RECORD_EXCEL)
                for _, row in df.iterrows():
                    dept   = str(row.get('Department', '')).strip()
                    school = str(row.get('School/Office', '')).strip()
                    emp    = str(row.get('Employee', '')).strip()
                    code   = str(row.get('Code', '')).strip()
                    if not (emp and code):
                        continue
                    vals = [dept, school, emp, code]
                    placeholders = ["?", "?", "?", "?"]
                    for i in range(1, 11):
                        off  = row.get(f'Receiving Office {i}')
                        rn   = row.get(f'Receiver Name {i}')
                        ts   = row.get(f'Timestamp {i}')
                        vals.append(None if pd.isna(off) else str(off))
                        vals.append(None if pd.isna(rn)  else str(rn))
                        vals.append(None if pd.isna(ts)  else str(ts))
                        placeholders.extend(["?", "?", "?"])
                    sql = ("INSERT OR IGNORE INTO routing_records "
                           "(department, school_office, employee, code, " +
                           ", ".join([f"receiving_office_{i}, receiver_name_{i}, timestamp_{i}" for i in range(1, 11)]) +
                           f") VALUES ({', '.join(placeholders)})")
                    cursor.execute(sql, vals)
                conn.commit()
                print("Imported existing routing_record.xlsx successfully.")
            except Exception as e:
                print(f"Error importing existing routing record excel: {e}")

        elif os.path.exists(OLD_LOG_EXCEL):
            try:
                df = pd.read_excel(OLD_LOG_EXCEL)
                df = df.sort_values(by='Timestamp')
                for _, row in df.iterrows():
                    code  = str(row.get('Code', '')).strip()
                    emp   = str(row.get('Employee Name', '')).strip()
                    dept  = str(row.get('Department', 'Unknown')).strip()
                    school = str(row.get('School/Office', 'Unknown')).strip()
                    ts    = str(row.get('Timestamp', datetime.now().strftime("%m/%d/%Y"))).strip()
                    if not (code and emp):
                        continue
                    cursor.execute("SELECT * FROM routing_records WHERE code = ? AND employee = ?", (code, emp))
                    exist = cursor.fetchone()
                    if not exist:
                        cursor.execute(
                            "INSERT INTO routing_records (department, school_office, employee, code, receiving_office_1, receiver_name_1, timestamp_1) "
                            "VALUES (?, ?, ?, ?, ?, ?, ?)",
                            (dept, school, emp, code, school, '', ts)
                        )
                    else:
                        record_id = exist['id']
                        for idx in range(1, 11):
                            if not exist[f'receiving_office_{idx}']:
                                cursor.execute(
                                    f"UPDATE routing_records SET receiving_office_{idx} = ?, timestamp_{idx} = ? WHERE id = ?",
                                    (school, ts, record_id)
                                )
                                break
                conn.commit()
                print("Migrated scan logs from scan_log.xlsx successfully.")
            except Exception as e:
                print(f"Error migrating scan logs: {e}")
    
    conn.close()
    trigger_excel_export()


# --- GOOGLE SHEETS INTEGRATION ---

# Cache for the gspread client to prevent repeated auth roundtrips (takes 1-2s per call)
_cached_gspread_client = None
_cached_creds_json = None
_gspread_lock = threading.Lock()

def _get_gspread_client(service_account_json):
    """Retrieve or initialize a cached gspread client instance."""
    global _cached_gspread_client, _cached_creds_json
    with _gspread_lock:
        if _cached_gspread_client is not None and _cached_creds_json == service_account_json:
            return _cached_gspread_client
        
        import gspread
        from google.oauth2.service_account import Credentials
        import json
        
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds_dict = json.loads(service_account_json)
        creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(creds)
        
        _cached_gspread_client = client
        _cached_creds_json = service_account_json
        return client


def test_google_sheets_connection(sheet_id, service_account_json):
    """Test connecting to a Google Sheet and initialize headers if empty."""
    import gspread
    try:
        client = _get_gspread_client(service_account_json)
        sh = client.open_by_key(sheet_id)
        
        try:
            ws = sh.worksheet("Routing Records")
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet(title="Routing Records", rows=1000, cols=50)
        
        # Headers include doc_type, receiver_name and remarks
        headers = ["Department", "School/Office", "Employee", "Code", "Doc Type", "Remarks"]
        for i in range(1, 11):
            headers += [f"Receiving Office {i}", f"Receiver Name {i}", f"Timestamp {i}"]
        
        existing_headers = ws.row_values(1)
        if not existing_headers:
            ws.insert_row(headers, 1)
            end_col = _col_to_letter(len(headers))
            ws.format(f"A1:{end_col}1", {
                "backgroundColor": {"red": 0.05, "green": 0.25, "blue": 0.45},
                "horizontalAlignment": "CENTER",
                "textFormat": {
                    "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                    "bold": True, "fontSize": 11
                }
            })
            ws.freeze(rows=1)
        
        return {"status": "success", "title": sh.title}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def _col_to_letter(col):
    """Convert 1-based column number to spreadsheet letter (e.g. 1→A, 27→AA)."""
    let = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        let = chr(65 + remainder) + let
    return let


def _gsheets_push_with_retry(push_fn, *args, max_retries=3, **kwargs):
    """Call push_fn(*args, **kwargs) with exponential-backoff retry on failure."""
    last_err = None
    for attempt in range(1, max_retries + 1):
        try:
            push_fn(*args, **kwargs)
            return True
        except Exception as e:
            last_err = e
            wait = 2 ** attempt  # 2s, 4s, 8s
            print(f"GSheets push attempt {attempt}/{max_retries} failed: {e}. Retrying in {wait}s...")
            time.sleep(wait)
    print(f"GSheets push permanently failed after {max_retries} attempts: {last_err}")
    return False


def push_row_to_gsheets(sheet_id, service_account_json, row_dict):
    """Safely updates or appends a routing row in Google Sheets."""
    import gspread
    client = _get_gspread_client(service_account_json)
    sh = client.open_by_key(sheet_id)
    ws = sh.worksheet("Routing Records")
    
    # Always use fixed 10 stages for consistent column alignment
    target_headers = ["Department", "School/Office", "Employee", "Code", "Doc Type", "Remarks", "Status", "Employee ID"]
    for i in range(1, 11):
        target_headers += [f"Receiving Office {i}", f"Receiver Name {i}", f"Timestamp {i}"]
    
    # Get doc_type from row_dict; if empty, look it up from code_lookup
    doc_type = row_dict.get('doc_type', '') or row_dict.get('doc_type_val', '') or ''
    if not doc_type and row_dict.get('code'):
        try:
            _conn = get_db_connection()
            _row = _conn.execute("SELECT doc_type FROM code_lookup WHERE code = ?", (row_dict['code'],)).fetchone()
            _conn.close()
            if _row:
                doc_type = _row['doc_type'] or ''
        except Exception:
            pass
    
    row_data = [
        row_dict.get('department', ''),
        row_dict.get('school_office', ''),
        row_dict.get('employee', ''),
        row_dict.get('code', ''),
        doc_type,
        row_dict.get('remarks', ''),
        row_dict.get('status', 'for signature'),
        row_dict.get('employee_id', '') or '',
    ]
    for i in range(1, 11):
        row_data.append(row_dict.get(f'receiving_office_{i}') or "")
        row_data.append(row_dict.get(f'receiver_name_{i}') or "")
        row_data.append(row_dict.get(f'timestamp_{i}') or "")
    
    # Expand headers if needed — use a lightweight header check
    try:
        existing_headers = ws.row_values(1)
    except Exception:
        existing_headers = []
    if len(existing_headers) < len(target_headers):
        end_col = _col_to_letter(len(target_headers))
        ws.update(range_name=f"A1:{end_col}1", values=[target_headers])
        ws.format(f"A1:{end_col}1", {
            "backgroundColor": {"red": 0.05, "green": 0.25, "blue": 0.45},
            "horizontalAlignment": "CENTER",
            "textFormat": {
                "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                "bold": True, "fontSize": 11
            }
        })
    
    # Find matching row by Code (column D = 4) — read only that column instead of the full sheet
    code_to_match = str(row_dict.get('code', '')).strip().upper()
    emp_to_match  = str(row_dict.get('employee', '')).strip().lower()
    match_row_num = None
    try:
        code_col = ws.col_values(4)  # Column D = Code (faster than get_all_values)
        for idx, cell_code in enumerate(code_col[1:], start=2):  # skip header
            if str(cell_code).strip().upper() == code_to_match:
                # Verify Employee (column C = 3) matches only for this candidate row
                emp_cell = ws.cell(idx, 3).value
                if str(emp_cell).strip().lower() == emp_to_match:
                    match_row_num = idx
                    break
    except Exception:
        # Fallback: read all values if col_values approach fails
        all_rows = ws.get_all_values()
        for idx, row in enumerate(all_rows[1:], start=2):
            if len(row) >= 4:
                if str(row[2]).strip().lower() == emp_to_match and str(row[3]).strip().upper() == code_to_match:
                    match_row_num = idx
                    break
    
    end_col = _col_to_letter(len(target_headers))
    if match_row_num:
        ws.update(range_name=f"A{match_row_num}:{end_col}{match_row_num}", values=[row_data])
        print(f"Updated Google Sheets row {match_row_num} for {row_dict.get('employee')}")
    else:
        ws.append_row(row_data)
        print(f"Appended new Google Sheets row for {row_dict.get('employee')}")


def batch_push_routing_rows(sheet_id, service_account_json, rows):
    """Push multiple routing records in a single batch operation.
    Reads the sheet once, then does one batch-update + one batch-append.
    """
    if not rows:
        return
    import gspread
    client = _get_gspread_client(service_account_json)
    sh = client.open_by_key(sheet_id)
    ws = sh.worksheet("Routing Records")

    target_headers = ["Department", "School/Office", "Employee", "Code", "Doc Type", "Remarks", "Status", "Employee ID"]
    for i in range(1, 11):
        target_headers += [f"Receiving Office {i}", f"Receiver Name {i}", f"Timestamp {i}"]
    end_col = _col_to_letter(len(target_headers))

    # Read current sheet data once
    all_values = ws.get_all_values()
    existing_headers = all_values[0] if all_values else []
    data_rows = all_values[1:] if len(all_values) > 1 else []

    # Build lookup: (Code, Employee) -> row number (1-based)
    existing = {}
    for idx, row in enumerate(data_rows):
        if len(row) >= 4:
            cell_code = str(row[3]).strip().upper()
            cell_emp  = str(row[2]).strip().lower()
            if cell_code and cell_emp:
                existing[(cell_code, cell_emp)] = idx + 2

    # Expand headers if needed
    if len(existing_headers) < len(target_headers):
        ws.update(range_name=f"A1:{end_col}1", values=[target_headers])

    # Pre-fetch doc_type for all codes
    codes_to_lookup = list({str(r.get('code', '')).strip() for r in rows if r.get('code')})
    code_doc_map = {}
    if codes_to_lookup:
        try:
            conn = get_db_connection()
            placeholders = ','.join(['?'] * len(codes_to_lookup))
            _rows = conn.execute(
                f"SELECT code, doc_type FROM code_lookup WHERE code IN ({placeholders})",
                codes_to_lookup
            ).fetchall()
            code_doc_map = {r['code']: r['doc_type'] or '' for r in _rows}
            conn.close()
        except Exception:
            pass

    # Separate rows into updates vs appends
    update_batch = []
    append_batch = []

    for r in rows:
        doc_type = r.get('doc_type', '') or code_doc_map.get(r.get('code', ''), '')
        row_data = [
            r.get('department', ''),
            r.get('school_office', ''),
            r.get('employee', ''),
            r.get('code', ''),
            doc_type,
            r.get('remarks', ''),
            r.get('status', 'for signature'),
            r.get('employee_id', '') or '',
        ]
        for i in range(1, 11):
            row_data.append(r.get(f'receiving_office_{i}') or "")
            row_data.append(r.get(f'receiver_name_{i}') or "")
            row_data.append(r.get(f'timestamp_{i}') or "")

        code = str(r.get('code', '')).strip().upper()
        emp  = str(r.get('employee', '')).strip().lower()
        key  = (code, emp)

        if key in existing:
            update_batch.append({
                'range': f'A{existing[key]}:{end_col}{existing[key]}',
                'values': [row_data]
            })
        else:
            append_batch.append(row_data)

    if update_batch:
        ws.batch_update(update_batch)
    if append_batch:
        ws.append_rows(append_batch, value_input_option='USER_ENTERED')

    print(f"Batch pushed {len(rows)} routing rows ({len(update_batch)} updates, {len(append_batch)} appends)")


def push_all_routing_to_gsheets(sheet_id, service_account_json, rows):
    """Batch write all routing records to Google Sheets in one fast operation."""
    import gspread
    client = _get_gspread_client(service_account_json)
    sh = client.open_by_key(sheet_id)
    try:
        ws = sh.worksheet("Routing Records")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="Routing Records", rows=max(1000, len(rows) + 100), cols=50)
    
    headers = ["Department", "School/Office", "Employee", "Code", "Doc Type", "Remarks", "Status", "Employee ID"]
    for i in range(1, 11):
        headers += [f"Receiving Office {i}", f"Receiver Name {i}", f"Timestamp {i}"]
    
    # Batch-fetch doc_type for all codes in one query (avoids N+1 lookups)
    all_codes = [dict(r).get('code', '') for r in rows if dict(r).get('code')]
    code_doc_type_map = {}
    if all_codes:
        try:
            _conn = get_db_connection()
            placeholders = ','.join(['?'] * len(all_codes))
            _rows = _conn.execute(f"SELECT code, doc_type FROM code_lookup WHERE code IN ({placeholders})", all_codes).fetchall()
            code_doc_type_map = {row['code']: row['doc_type'] or '' for row in _rows}
            _conn.close()
        except Exception:
            pass
    
    rows_to_write = [headers]
    for r in rows:
        row_dict = dict(r)
        doc_type = row_dict.get('doc_type', '') or row_dict.get('doc_type_val', '') or ''
        if not doc_type and row_dict.get('code'):
            doc_type = code_doc_type_map.get(row_dict['code'], '')
        row_data = [
            row_dict.get('department', ''),
            row_dict.get('school_office', ''),
            row_dict.get('employee', ''),
            row_dict.get('code', ''),
            doc_type,
            row_dict.get('remarks', ''),
            row_dict.get('status', 'for signature'),
            row_dict.get('employee_id', '') or '',
        ]
        for i in range(1, 11):
            row_data.append(row_dict.get(f'receiving_office_{i}') or "")
            row_data.append(row_dict.get(f'receiver_name_{i}') or "")
            row_data.append(row_dict.get(f'timestamp_{i}') or "")
        rows_to_write.append(row_data)
        
    ws.clear()
    end_col = _col_to_letter(len(headers))
    ws.update(range_name=f"A1:{end_col}{len(rows_to_write)}", values=rows_to_write)
    ws.format(f"A1:{end_col}1", {
        "backgroundColor": {"red": 0.05, "green": 0.25, "blue": 0.45},
        "horizontalAlignment": "CENTER",
        "textFormat": {
            "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
            "bold": True, "fontSize": 11
        }
    })
    ws.freeze(rows=1)


def push_all_codes_to_gsheets(sheet_id, service_account_json, rows):
    """Batch write all code lookups to Google Sheets in one fast operation."""
    import gspread
    client = _get_gspread_client(service_account_json)
    sh = client.open_by_key(sheet_id)
    try:
        ws = sh.worksheet("Code Lookup")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="Code Lookup", rows=max(1000, len(rows) + 100), cols=10)
    
    headers = ["Code", "Department", "School/Office", "Employees", "Doc Type", "Generated"]
    rows_to_write = [headers]
    for r in rows:
        row_dict = dict(r)
        rows_to_write.append([
            row_dict.get('code', ''),
            row_dict.get('department', ''),
            row_dict.get('school_office', ''),
            row_dict.get('employees', ''),
            row_dict.get('doc_type', ''),
            row_dict.get('generated_at', '')
        ])
        
    ws.clear()
    ws.update(range_name=f"A1:F{len(rows_to_write)}", values=rows_to_write)
    ws.format("A1:F1", {"backgroundColor": {"red": 0.05, "green": 0.25, "blue": 0.45}, "textFormat": {"foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}, "bold": True}})
    ws.freeze(rows=1)


def push_all_users_to_gsheets(sheet_id, service_account_json, rows):
    """Batch write all users to Google Sheets in one fast operation."""
    import gspread
    client = _get_gspread_client(service_account_json)
    sh = client.open_by_key(sheet_id)
    try:
        ws = sh.worksheet("Users")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="Users", rows=max(100, len(rows) + 10), cols=10)
    
    headers = ["Username", "Password Hash", "Role", "Status", "School/Office", "Email", "Requires Password Change", "Supervised Schools", "Employee Name"]
    rows_to_write = [headers]
    for r in rows:
        row_dict = dict(r)
        rows_to_write.append([
            row_dict.get('username', ''),
            row_dict.get('password_hash', ''),
            row_dict.get('role', ''),
            row_dict.get('status', ''),
            row_dict.get('school_office', ''),
            row_dict.get('email', ''),
            row_dict.get('requires_password_change', 0),
            row_dict.get('supervised_schools', ''),
            row_dict.get('employee_name', ''),
        ])
        
    ws.clear()
    end_col = 'I'
    ws.update(range_name=f"A1:{end_col}{len(rows_to_write)}", values=rows_to_write)
    ws.format(f"A1:{end_col}1", {"backgroundColor": {"red": 0.05, "green": 0.25, "blue": 0.45}, "textFormat": {"foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}, "bold": True}})
    ws.freeze(rows=1)


def start_gsheets_worker():
    """Start background queue consumer for Google Sheets sync.
    Uses persisted sync_queue table (crash-safe) with read-once-per-table batching.
    No GSheets API calls if the queue is empty."""
    def _worker():
        print("Google Sheets Background Worker Started (persisted queue).")
        while True:
            try:
                # Wait for signal or poll every 5 seconds
                _sync_needed.wait(timeout=5)
                _sync_needed.clear()

                # In maintenance mode, skip all background pushes
                if get_setting('maintenance_mode') == 'True':
                    continue

                enabled  = get_setting('gsheets_enabled') == 'True'
                sheet_id = get_setting('gsheets_id')
                creds    = get_decrypted_setting('gsheets_credentials')
                if not enabled or not sheet_id or not creds:
                    continue

                # Process all pending sync items — read-once-per-table
                _process_pending_sync(sheet_id, creds)

            except Exception as ex:
                print(f"Worker critical error: {ex}")
                time.sleep(5)

    worker_thread = threading.Thread(target=_worker, daemon=True)
    worker_thread.start()


def _process_pending_sync(sheet_id, creds):
    """Read all pending items from sync_queue, process per table.
    Each table's batch uses a single read-write cycle to minimize API calls.
    This is a module-level function so it can also be called from the shutdown handler."""
    conn = get_db_connection()
    try:
        tables = conn.execute(
            "SELECT DISTINCT table_name FROM sync_queue WHERE status = 'pending'"
        ).fetchall()

        for (table_name,) in tables:
            items = conn.execute(
                "SELECT * FROM sync_queue WHERE table_name = ? AND status = 'pending' ORDER BY id LIMIT 500",
                (table_name,)
            ).fetchall()

            if not items:
                continue

            ids = tuple(item['id'] for item in items)
            id_placeholders = ','.join(['?'] * len(ids))

            # Mark as processing
            conn.execute(
                f"UPDATE sync_queue SET status = 'processing' WHERE id IN ({id_placeholders})",
                ids
            )
            conn.commit()

            try:
                if table_name == 'routing_records':
                    ref_ids = [item['ref_id'] for item in items if item['ref_id'] and item['operation'] != 'delete']
                    if ref_ids:
                        rid_ph = ','.join(['?'] * len(ref_ids))
                        rows = conn.execute(
                            f"SELECT * FROM routing_records WHERE id IN ({rid_ph})",
                            ref_ids
                        ).fetchall()
                        if rows:
                            _gsheets_push_with_retry(batch_push_routing_rows, sheet_id, creds, [dict(r) for r in rows])

                elif table_name == 'code_lookup':
                    # Full snapshot: push all codes to GSheets
                    if any(item['operation'] != 'delete' for item in items):
                        all_codes = conn.execute("SELECT * FROM code_lookup").fetchall()
                        if all_codes:
                            _gsheets_push_with_retry(push_all_codes_to_gsheets, sheet_id, creds, all_codes)

                elif table_name == 'users':
                    # Full snapshot: push all users to GSheets
                    if any(item['operation'] != 'delete' for item in items):
                        all_users = conn.execute("SELECT * FROM users").fetchall()
                        if all_users:
                            _gsheets_push_with_retry(push_all_users_to_gsheets, sheet_id, creds, all_users)

                # Handle deletions
                for item in items:
                    if item['operation'] == 'delete' and item['table_name'] == 'routing_records':
                        ref_id = item['ref_id']
                        if ref_id:
                            # Push a "deleted" marker to GSheets
                            row = conn.execute("SELECT * FROM routing_records WHERE id = ?", (ref_id,)).fetchone()
                            if row:
                                row_dict = dict(row)
                                row_dict['status'] = 'deleted'
                                _gsheets_push_with_retry(push_row_to_gsheets, sheet_id, creds, row_dict)

                # Mark all as done
                conn.execute(
                    f"UPDATE sync_queue SET status = 'done' WHERE id IN ({id_placeholders})",
                    ids
                )
                conn.commit()

            except Exception as e:
                print(f"Worker: error processing {table_name}: {e}")
                for item in items:
                    if item['retry_count'] < 3:
                        conn.execute(
                            "UPDATE sync_queue SET retry_count = retry_count + 1, last_error = ?, status = 'pending' WHERE id = ?",
                            (str(e)[:200], item['id'])
                        )
                    else:
                        conn.execute(
                            "UPDATE sync_queue SET retry_count = retry_count + 1, last_error = ?, status = 'failed' WHERE id = ?",
                            (str(e)[:200], item['id'])
                        )
                conn.commit()
    finally:
        conn.close()


def start_periodic_gsheets_pull():
    """Start a background daemon thread that calls pull_all_from_gsheets()
    on a configurable interval (default: 15 minutes, minimum: 1 minute).
    The interval is read from the 'gsheets_pull_interval' setting at the
    start of each cycle so it can be changed at runtime without a restart.
    Uses a short-sleep loop so the timer can be reset externally (e.g. after
    maintenance mode ends) without waiting for the full interval to elapse.
    """
    DEFAULT_INTERVAL = 1

    def _worker():
        global _periodic_last_run
        while True:
            try:
                raw = get_setting('gsheets_pull_interval', str(DEFAULT_INTERVAL))
                minutes = max(1, int(raw))
            except (ValueError, TypeError):
                minutes = DEFAULT_INTERVAL

            # In maintenance mode, skip pull but keep looping (timer does not advance)
            if get_setting('maintenance_mode') == 'True':
                update_last_pull_info("paused", "Maintenance mode active — periodic pull paused.")
                time.sleep(5)
                continue

            # Check if enough time has elapsed since the last pull
            if time.time() - _periodic_last_run >= minutes * 60:
                try:
                    update_last_pull_info("syncing", "Refreshing SQLite from Google Sheets...")
                    print(f"[Periodic Pull] Refreshing SQLite from Google Sheets...")
                    count, errors = pull_all_from_gsheets()
                    if errors:
                        for e in errors:
                            print(f"[Periodic Pull] Warning: {e}")
                    if count is not None:
                        update_last_pull_info("ok", f"Refreshed {count} routing records from Google Sheets.")
                        print(f"[Periodic Pull] OK — {count} routing records refreshed.")
                    else:
                        update_last_pull_info("ok", "Periodic pull completed.")
                        print(f"[Periodic Pull] OK (no routing records count).")
                except Exception as e:
                    update_last_pull_info("error", f"Periodic pull failed: {e}")
                    print(f"[Periodic Pull] Error: {e}")
                    time.sleep(30)

                _periodic_last_run = time.time()

            time.sleep(5)

    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    print(f"Periodic GSheets pull worker started (interval: {DEFAULT_INTERVAL} min, configurable via 'gsheets_pull_interval' setting).")


def reset_periodic_pull_timer():
    """Reset the periodic pull countdown so the next automatic pull happens
    after a full interval (default 15 min) from now. Used when exiting
    maintenance mode so the timer starts fresh."""
    global _periodic_last_run
    _periodic_last_run = time.time()


def enqueue_sync(table_name, ref_id=None, code=None, employee=None, operation='upsert'):
    """Persist a sync task to the sync_queue table. The background worker
    picks it up on its next cycle. Returns immediately (no API calls).
    SQLite write + queue enqueue in same transaction for crash safety."""
    try:
        conn = get_db_connection()
        conn.execute(
            "INSERT INTO sync_queue (table_name, ref_id, code, employee, operation, created_at, status) "
            "VALUES (?, ?, ?, ?, ?, ?, 'pending')",
            (table_name, ref_id, code, employee, operation, datetime.now().isoformat())
        )
        conn.commit()
        conn.close()
        _sync_needed.set()
    except Exception as e:
        print(f"enqueue_sync error: {e}")


def enqueue_sync_batch(table_name, ref_ids, operation='upsert'):
    """Enqueue multiple sync tasks in one shot."""
    if not ref_ids:
        return
    try:
        conn = get_db_connection()
        now = datetime.now().isoformat()
        for rid in ref_ids:
            conn.execute(
                "INSERT INTO sync_queue (table_name, ref_id, operation, created_at, status) "
                "VALUES (?, ?, ?, ?, 'pending')",
                (table_name, rid, operation, now)
            )
        conn.commit()
        conn.close()
        _sync_needed.set()
    except Exception as e:
        print(f"enqueue_sync_batch error: {e}")


def queue_sync(row_id):
    """Queue a single row update — legacy wrapper."""
    enqueue_sync('routing_records', ref_id=row_id)


def queue_sync_batch(row_ids):
    """Queue a batch of row updates — legacy wrapper."""
    enqueue_sync_batch('routing_records', row_ids)


def bulk_sync_to_gsheets():
    """Push ALL existing data (routing records, codes, users, master_data) to Google Sheets.
    Returns (count, error_messages)."""
    enabled = get_setting('gsheets_enabled') == 'True'
    sheet_id = get_setting('gsheets_id')
    creds = get_decrypted_setting('gsheets_credentials')
    
    if not enabled or not sheet_id or not creds:
        return 0, ["Google Sheets sync is not configured. Enable it and save credentials first."]
    
    errors = []
    count = 0
    
    try:
        conn = get_db_connection()
        
        # 1. Bulk push routing records in a single batch operation
        routing_rows = conn.execute("SELECT * FROM routing_records").fetchall()
        try:
            _gsheets_push_with_retry(push_all_routing_to_gsheets, sheet_id, creds, routing_rows)
            count += len(routing_rows)
        except Exception as e:
            errors.append(f"RoutingRecords: {str(e)[:100]}")
        
        # 2. Bulk push code lookups in a single batch operation
        code_rows = conn.execute("SELECT * FROM code_lookup").fetchall()
        try:
            _gsheets_push_with_retry(push_all_codes_to_gsheets, sheet_id, creds, code_rows)
        except Exception as e:
            errors.append(f"CodeLookup: {str(e)[:100]}")
                
        # 3. Bulk push users in a single batch operation
        user_rows = conn.execute("SELECT * FROM users").fetchall()
        try:
            _gsheets_push_with_retry(push_all_users_to_gsheets, sheet_id, creds, user_rows)
        except Exception as e:
            errors.append(f"Users: {str(e)[:100]}")

        # 4. Push raw master_db.xlsx sheets to Google Sheets (3 separate worksheets)
        try:
            _gsheets_push_with_retry(push_master_db_to_gsheets, sheet_id, creds)
        except Exception as e:
            errors.append(f"MasterData: {str(e)[:100]}")

        # 5. Push new RDBMS sheets (Schools, Employees — full snapshot)
        try:
            _gsheets_push_with_retry(push_schools_to_gsheets, sheet_id, creds)
        except Exception as e:
            errors.append(f"Schools: {str(e)[:100]}")
        try:
            _gsheets_push_with_retry(push_employees_to_gsheets, sheet_id, creds)
        except Exception as e:
            errors.append(f"Employees: {str(e)[:100]}")

        # 6. Push Code Employees (append-only)
        try:
            _gsheets_push_with_retry(push_code_employees_to_gsheets, sheet_id, creds)
        except Exception as e:
            errors.append(f"CodeEmployees: {str(e)[:100]}")

        conn.close()
        return count, errors
    except Exception as e:
        return 0, [str(e)]

def sync_master_to_employees():
    """Backfill schools, employees, code_employees from master_data + code_lookup.
    
    1. Upsert schools from master_data (distinct school_office + department)
    2. Upsert employees from master_data (employee_name + school_id)
    3. Create code_employees links from code_lookup.employees ||| split
    """
    conn = get_db_connection()
    cursor = conn.cursor()
    
    school_count = 0
    emp_count = 0
    ce_count = 0
    rr_count = 0
    
    # 1. Upsert schools
    schools = cursor.execute(
        "SELECT DISTINCT school_office, department FROM master_data"
    ).fetchall()
    for row in schools:
        cursor.execute(
            "INSERT OR IGNORE INTO schools (name, department) VALUES (?, ?)",
            (row['school_office'], row['department'])
        )
        if cursor.rowcount > 0:
            school_count += 1
    
    # 2. Upsert employees
    employees = cursor.execute(
        "SELECT md.employee_name, md.school_office FROM master_data md"
    ).fetchall()
    for row in employees:
        s = cursor.execute(
            "SELECT id FROM schools WHERE name = ?", (row['school_office'],)
        ).fetchone()
        if s:
            cursor.execute(
                "INSERT OR IGNORE INTO employees (name, school_id) VALUES (?, ?)",
                (row['employee_name'], s['id'])
            )
            if cursor.rowcount > 0:
                emp_count += 1
    
    # 3. Create code_employees links from code_lookup ||| split
    codes = cursor.execute("SELECT code, employees, school_office FROM code_lookup").fetchall()
    for code_row in codes:
        code = code_row['code']
        emp_names = code_row['employees'].split('|||') if code_row['employees'] else []
        school = code_row['school_office']
        s = cursor.execute("SELECT id FROM schools WHERE name = ?", (school,)).fetchone()
        if not s:
            continue
        for emp_name in emp_names:
            emp_name = emp_name.strip()
            if not emp_name:
                continue
            e = cursor.execute(
                "SELECT id FROM employees WHERE name = ? AND school_id = ?",
                (emp_name, s['id'])
            ).fetchone()
            if not e:
                # Employee not in master_data — create on the fly
                cursor.execute(
                    "INSERT OR IGNORE INTO employees (name, school_id) VALUES (?, ?)",
                    (emp_name, s['id'])
                )
                if cursor.rowcount > 0:
                    emp_count += 1
                e = cursor.execute(
                    "SELECT id FROM employees WHERE name = ? AND school_id = ?",
                    (emp_name, s['id'])
                ).fetchone()
            if e:
                cursor.execute(
                    "INSERT OR IGNORE INTO code_employees (code, employee_id) VALUES (?, ?)",
                    (code, e['id'])
                )
                if cursor.rowcount > 0:
                    ce_count += 1
    
    # 4. Link routing_records.employee_id
    rrs = cursor.execute("SELECT id, employee, school_office FROM routing_records WHERE employee_id IS NULL").fetchall()
    for rr in rrs:
        e = cursor.execute("""
            SELECT e.id FROM employees e
            JOIN schools s ON s.id = e.school_id
            WHERE e.name = ? AND s.name = ?
        """, (rr['employee'], rr['school_office'])).fetchone()
        if e:
            cursor.execute("UPDATE routing_records SET employee_id = ? WHERE id = ?", (e['id'], rr['id']))
            rr_count += 1
    
    conn.commit()
    conn.close()
    print(f"sync_master_to_employees: {school_count} schools, {emp_count} employees, {ce_count} code_employees, {rr_count} routing_records linked.")
    return school_count, emp_count, ce_count, rr_count


def force_import_from_excel():
    """Force reload data from Excel without exporting back - preserves manual Excel edits."""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    imported_count = 0
    
    # 1. Import code_lookup
    if os.path.exists(CODE_LOOKUP_EXCEL):
        try:
            df = pd.read_excel(CODE_LOOKUP_EXCEL)
            for _, row in df.iterrows():
                code = str(row.get('Code', '')).strip()
                if code:
                    # Check if code already exists in SQLite
                    cursor.execute("SELECT code FROM code_lookup WHERE code = ?", (code,))
                    if not cursor.fetchone():
                        dept = str(row.get('Department', 'Unknown')).strip()
                        school = str(row.get('School/Office', 'Unknown')).strip()
                        employees = str(row.get('Employees', '')).strip()
                        doc_type = str(row.get('Doc Type', '')).strip()
                        gen = str(row.get('Generated', datetime.now().strftime("%m/%d/%Y"))).strip()
                        cursor.execute(
                            "INSERT INTO code_lookup (code, department, school_office, employees, doc_type, generated_at) "
                            "VALUES (?, ?, ?, ?, ?, ?)",
                            (code, dept, school, employees, doc_type, gen)
                        )
                        imported_count += 1
            conn.commit()
            print(f"Force imported {imported_count} code lookup records.")
        except Exception as e:
            print(f"Error force importing code lookup: {e}")
    
    # 2. Import routing_records
    if os.path.exists(ROUTING_RECORD_EXCEL):
        try:
            df = pd.read_excel(ROUTING_RECORD_EXCEL)
            for _, row in df.iterrows():
                code = str(row.get('Code', '')).strip()
                emp = str(row.get('Employee', '')).strip()
                if code and emp:
                    # Check if record already exists in SQLite
                    cursor.execute("SELECT id FROM routing_records WHERE code = ? AND employee = ?", (code, emp))
                    if not cursor.fetchone():
                        dept = str(row.get('Department', '')).strip()
                        school = str(row.get('School/Office', '')).strip()
                        # Get all receiving_office columns dynamically
                        vals = [dept, school, emp, code]
                        for col in df.columns:
                            if 'Receiving Office' in col or 'Timestamp' in col or 'Receiver Name' in col:
                                vals.append(str(row.get(col, '')))
                        placeholders = ['?'] * len(vals)
                        col_names = ['department', 'school_office', 'employee', 'code']
                        for col in df.columns:
                            if 'Receiving Office' in col:
                                col_names.append(col)
                            elif 'Timestamp' in col:
                                col_names.append(col)
                            elif 'Receiver Name' in col:
                                col_names.append(col)
                        sql = f"INSERT INTO routing_records ({', '.join(col_names)}) VALUES ({', '.join(placeholders)})"
                        cursor.execute(sql, vals)
                        imported_count += 1
            conn.commit()
            print(f"Force imported routing records.")
        except Exception as e:
            print(f"Error force importing routing records: {e}")
    
    conn.close()
    return imported_count
def _pull_master_data_from_gsheets(sh, cursor):
    """Pull the 3 raw master sheets from GSheets and rebuild the SQLite master_data table.
    
    Sheet1 → school→department mapping + Sheet1 employees
    Sheet3 → Senior High Schools list (used for Sheet2 dept inference)
    Sheet2 → Additional employees (name constructed from FIRSTNAME/MIDDLE INITIAL/LASTNAME/SUFFIX)
    
    Falls back to the local master_db.xlsx if any worksheet is missing.
    """
    # Check which GSheets worksheets exist
    ws_names = {ws.title for ws in sh.worksheets()}
    has_sheet1 = "Master Data Sheet1" in ws_names
    has_sheet2 = "Master Data Sheet2" in ws_names
    has_sheet3 = "Master Data Sheet3" in ws_names
    
    use_gsheets = has_sheet1 or has_sheet2 or has_sheet3
    
    if not use_gsheets:
        print("GSheets pull: no 'Master Data Sheet*' worksheets found — seeding from local Excel if available.")
        _seed_master_data_from_excel(cursor)
        return
    
    # ── Wipe master_data before rebuilding ──
    cursor.execute("DELETE FROM master_data")
    
    school_dept_map = {}
    shs_schools = set()
    sheet1_count = 0
    sheet2_count = 0
    
    # ── 1. Pull Sheet1 ──────────────────────────────────────────────────────
    if has_sheet1:
        try:
            ws1 = sh.worksheet("Master Data Sheet1")
            records = ws1.get_all_records()
            for row in records:
                dept = str(row.get('Department', '')).strip()
                school = str(row.get('School/Office', '')).strip()
                emp = str(row.get('Employee_Name', '')).strip()
                if dept and school and emp:
                    cursor.execute(
                        "INSERT OR IGNORE INTO master_data (department, school_office, employee_name) VALUES (?, ?, ?)",
                        (dept, school, emp)
                    )
                    if cursor.rowcount > 0:
                        sheet1_count += 1
                    school_key = school.lower().strip()
                    if school_key not in school_dept_map:
                        school_dept_map[school_key] = dept
            print(f"GSheets pull: Sheet1 — {sheet1_count} entries restored.")
        except Exception as e:
            print(f"GSheets pull: error reading 'Master Data Sheet1': {e}")
    else:
        # Fallback: try reading Sheet1 from local Excel
        _seed_master_data_sheet1_from_excel(cursor, school_dept_map)
    
    # ── 2. Pull Sheet3 (SHS schools list) ───────────────────────────────────
    if has_sheet3:
        try:
            ws3 = sh.worksheet("Master Data Sheet3")
            records = ws3.get_all_records()
            # get_all_records uses first row as headers; for Sheet3 we need all values
            # Sheet3 may have no standard headers — read raw values
            all_vals = ws3.get_all_values()
            for row in all_vals:
                for val in row:
                    val_str = str(val).strip()
                    if val_str and val_str.lower() != 'nan':
                        shs_schools.add(val_str.lower())
            print(f"GSheets pull: Sheet3 — {len(shs_schools)} SHS school entries loaded.")
        except Exception as e:
            print(f"GSheets pull: error reading 'Master Data Sheet3': {e}")
    else:
        # Fallback: read from local Excel
        _load_shs_schools_from_excel(shs_schools)
    
    # ── 3. Pull Sheet2 ──────────────────────────────────────────────────────
    if has_sheet2:
        try:
            ws2 = sh.worksheet("Master Data Sheet2")
            records = ws2.get_all_records()
            for row in records:
                # Prefer pre-built EMPLOYEE_NAME column (ArrayFormula on Sheet2)
                emp_raw = row.get('EMPLOYEE_NAME', '')
                if emp_raw and str(emp_raw).strip() and str(emp_raw).strip().lower() != 'nan':
                    emp = str(emp_raw).strip()
                else:
                    # Fallback: construct from individual name columns
                    name_parts = [
                        row.get('FIRSTNAME'),
                        row.get('MIDDLE INITIAL'),
                        row.get('LASTNAME'),
                        row.get('SUFFIX')
                    ]
                    emp = " ".join([str(n).strip() for n in name_parts if n and str(n).strip() and str(n).strip().lower() != 'nan'])
                school = str(row.get('SCHOOL/ OFFICE', '')).strip()
                
                if not emp or not school:
                    continue
                
                # Determine department
                school_key = school.lower().strip()
                dept = school_dept_map.get(school_key)
                
                if not dept:
                    if school_key in shs_schools or '(shs)' in school_key:
                        dept = 'SENIOR HIGH SCHOOL'
                    elif 'elementary' in school_key or ' es' in school_key:
                        dept = 'ELEMENTARY SCHOOL'
                    elif 'high school' in school_key or ' hs' in school_key or 'integrated' in school_key:
                        dept = 'SECONDARY SCHOOL'
                    else:
                        dept = 'SDO - MANILA'
                
                cursor.execute(
                    "INSERT OR IGNORE INTO master_data (department, school_office, employee_name) VALUES (?, ?, ?)",
                    (dept, school, emp)
                )
                if cursor.rowcount > 0:
                    sheet2_count += 1
            print(f"GSheets pull: Sheet2 — {sheet2_count} entries restored.")
        except Exception as e:
            print(f"GSheets pull: error reading 'Master Data Sheet2': {e}")
    else:
        # Fallback: try reading Sheet2 from local Excel
        _seed_master_data_sheet2_from_excel(cursor, school_dept_map, shs_schools)
    
    cursor.execute("SELECT COUNT(*) FROM master_data")
    total = cursor.fetchone()[0]
    print(f"GSheets pull: master data total — {total} entries.")


def _seed_master_data_from_excel(cursor):
    """Seed master_data entirely from the local master_db.xlsx (fallback when no GSheets sheets exist)."""
    if not os.path.exists(MASTER_DB):
        print("GSheets pull fallback: master_db.xlsx not found — no master data seeded.")
        return
    
    try:
        xl = pd.ExcelFile(MASTER_DB)
        school_dept_map = {}
        
        # Sheet1
        if 'Sheet1' in xl.sheet_names:
            df1 = pd.read_excel(xl, 'Sheet1')
            if all(col in df1.columns for col in ['Department', 'School/Office', 'Employee_Name']):
                for _, row in df1.iterrows():
                    dept = str(row.get('Department', '')).strip()
                    school = str(row.get('School/Office', '')).strip()
                    emp = str(row.get('Employee_Name', '')).strip()
                    if dept and school and emp:
                        cursor.execute(
                            "INSERT OR IGNORE INTO master_data (department, school_office, employee_name) VALUES (?, ?, ?)",
                            (dept, school, emp)
                        )
                        school_key = school.lower().strip()
                        if school_key not in school_dept_map:
                            school_dept_map[school_key] = dept
        
        # Sheet3
        shs_schools = set()
        if 'Sheet3' in xl.sheet_names:
            df3 = pd.read_excel(xl, 'Sheet3')
            for col in df3.columns:
                for val in df3[col].dropna():
                    val_str = str(val).strip()
                    if val_str:
                        shs_schools.add(val_str.lower())
        
        # Sheet2
        if 'Sheet2' in xl.sheet_names:
            df2 = pd.read_excel(xl, 'Sheet2')
            required_cols = ['FIRSTNAME', 'LASTNAME', 'SCHOOL/ OFFICE']
            if all(col in df2.columns for col in required_cols):
                for _, row in df2.iterrows():
                    name_parts = [row.get('FIRSTNAME'), row.get('MIDDLE INITIAL'), row.get('LASTNAME'), row.get('SUFFIX')]
                    emp = " ".join([str(n).strip() for n in name_parts if pd.notna(n) and str(n).strip()])
                    school = str(row.get('SCHOOL/ OFFICE', '')).strip()
                    if not emp or not school:
                        continue
                    school_key = school.lower().strip()
                    dept = school_dept_map.get(school_key)
                    if not dept:
                        if school_key in shs_schools or '(shs)' in school_key:
                            dept = 'SENIOR HIGH SCHOOL'
                        elif 'elementary' in school_key or ' es' in school_key:
                            dept = 'ELEMENTARY SCHOOL'
                        elif 'high school' in school_key or ' hs' in school_key or 'integrated' in school_key:
                            dept = 'SECONDARY SCHOOL'
                        else:
                            dept = 'SDO - MANILA'
                    cursor.execute(
                        "INSERT OR IGNORE INTO master_data (department, school_office, employee_name) VALUES (?, ?, ?)",
                        (dept, school, emp)
                    )
        
        cursor.execute("SELECT COUNT(*) FROM master_data")
        total = cursor.fetchone()[0]
        print(f"GSheets pull fallback: seeded {total} master_data entries from local Excel.")
    except Exception as e:
        print(f"GSheets pull fallback: error seeding from Excel: {e}")


def _seed_master_data_sheet1_from_excel(cursor, school_dept_map):
    """Seed Sheet1 data from local Excel into master_data (fallback)."""
    if not os.path.exists(MASTER_DB):
        return
    try:
        xl = pd.ExcelFile(MASTER_DB)
        if 'Sheet1' in xl.sheet_names:
            df1 = pd.read_excel(xl, 'Sheet1')
            if all(col in df1.columns for col in ['Department', 'School/Office', 'Employee_Name']):
                count = 0
                for _, row in df1.iterrows():
                    dept = str(row.get('Department', '')).strip()
                    school = str(row.get('School/Office', '')).strip()
                    emp = str(row.get('Employee_Name', '')).strip()
                    if dept and school and emp:
                        cursor.execute(
                            "INSERT OR IGNORE INTO master_data (department, school_office, employee_name) VALUES (?, ?, ?)",
                            (dept, school, emp)
                        )
                        if cursor.rowcount > 0:
                            count += 1
                        school_key = school.lower().strip()
                        if school_key not in school_dept_map:
                            school_dept_map[school_key] = dept
                print(f"GSheets pull fallback: seeded {count} Sheet1 entries from local Excel.")
    except Exception as e:
        print(f"GSheets pull fallback: error reading Sheet1 from Excel: {e}")


def _seed_master_data_sheet2_from_excel(cursor, school_dept_map, shs_schools):
    """Seed Sheet2 data from local Excel into master_data (fallback)."""
    if not os.path.exists(MASTER_DB):
        return
    try:
        xl = pd.ExcelFile(MASTER_DB)
        if 'Sheet2' in xl.sheet_names:
            df2 = pd.read_excel(xl, 'Sheet2')
            required_cols = ['FIRSTNAME', 'LASTNAME', 'SCHOOL/ OFFICE']
            if all(col in df2.columns for col in required_cols):
                count = 0
                for _, row in df2.iterrows():
                    name_parts = [row.get('FIRSTNAME'), row.get('MIDDLE INITIAL'), row.get('LASTNAME'), row.get('SUFFIX')]
                    emp = " ".join([str(n).strip() for n in name_parts if pd.notna(n) and str(n).strip()])
                    school = str(row.get('SCHOOL/ OFFICE', '')).strip()
                    if not emp or not school:
                        continue
                    school_key = school.lower().strip()
                    dept = school_dept_map.get(school_key)
                    if not dept:
                        if school_key in shs_schools or '(shs)' in school_key:
                            dept = 'SENIOR HIGH SCHOOL'
                        elif 'elementary' in school_key or ' es' in school_key:
                            dept = 'ELEMENTARY SCHOOL'
                        elif 'high school' in school_key or ' hs' in school_key or 'integrated' in school_key:
                            dept = 'SECONDARY SCHOOL'
                        else:
                            dept = 'SDO - MANILA'
                    cursor.execute(
                        "INSERT OR IGNORE INTO master_data (department, school_office, employee_name) VALUES (?, ?, ?)",
                        (dept, school, emp)
                    )
                    if cursor.rowcount > 0:
                        count += 1
                print(f"GSheets pull fallback: seeded {count} Sheet2 entries from local Excel.")
    except Exception as e:
        print(f"GSheets pull fallback: error reading Sheet2 from Excel: {e}")


def _load_shs_schools_from_excel(shs_schools):
    """Load SHS school names from local Excel Sheet3 (fallback)."""
    if not os.path.exists(MASTER_DB):
        return
    try:
        xl = pd.ExcelFile(MASTER_DB)
        if 'Sheet3' in xl.sheet_names:
            df3 = pd.read_excel(xl, 'Sheet3')
            for col in df3.columns:
                for val in df3[col].dropna():
                    val_str = str(val).strip()
                    if val_str:
                        shs_schools.add(val_str.lower())
    except Exception:
        pass


# Cross-platform file lock for DB sync operations.
# Prevents multiple gunicorn workers from running pull_all_from_gsheets() concurrently.
_LOCK_FILE = os.path.join(os.path.dirname(DB_FILE) or '.', '.gsheets_sync.lock')

def _acquire_sync_lock():
    fd = None
    try:
        fd = os.open(_LOCK_FILE, os.O_CREAT | os.O_RDWR | os.O_TRUNC)
        if os.name == 'nt':
            import msvcrt
            msvcrt.locking(fd, msvcrt.LK_NBLCK, 1)
        else:
            import fcntl
            fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
        return fd
    except (IOError, OSError, ImportError, BlockingIOError):
        if fd is not None:
            try: os.close(fd)
            except: pass
        return None

def _release_sync_lock(fd):
    if fd is None:
        return
    try:
        if os.name == 'nt':
            import msvcrt
            msvcrt.locking(fd, msvcrt.LK_UNLCK, 1)
        else:
            import fcntl
            fcntl.flock(fd, fcntl.LOCK_UN)
        os.close(fd)
    except Exception:
        pass
    try:
        os.remove(_LOCK_FILE)
    except Exception:
        pass

def pull_all_from_gsheets(sheet_id=None, creds=None):
    """Upsert all tables from Google Sheets into SQLite (one-way pull).
    Uses INSERT ... ON CONFLICT DO UPDATE — never deletes local-only rows.
    Uses a cross-process file lock so only one gunicorn worker pulls at a time.
    Returns (routing_record_count, error_list)."""
    if not sheet_id or not creds:
        enabled = get_setting('gsheets_enabled') == 'True'
        if not enabled: return 0, ["GSheets sync disabled."]
        sheet_id = get_setting('gsheets_id')
        creds = get_decrypted_setting('gsheets_credentials')
        if not sheet_id or not creds: return 0, ["Missing GSheets credentials."]
    
    # Only one gunicorn worker pulls at a time; skip if another worker is already syncing
    lock_fd = _acquire_sync_lock()
    if lock_fd is None:
        print("[GSheets pull] Another worker is already syncing — skipping this pull.")
        return 0, ["Sync already in progress by another worker."]
        
    import gspread
    from google.oauth2.service_account import Credentials
    import json
    
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds_dict = json.loads(creds)
        credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(credentials)
        sh = client.open_by_key(sheet_id)
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        count = 0
        
        # ── 1. Pull Routing Records (upsert only — preserves local-only rows) ──
        try:
            ws_routing = sh.worksheet("Routing Records")
            records = ws_routing.get_all_records()
            for row in records:
                code = str(row.get('Code', '')).strip()
                emp = str(row.get('Employee', '')).strip()
                if not code or not emp: continue
                
                dept = str(row.get('Department', '')).strip()
                school = str(row.get('School/Office', '')).strip()
                doc_type = str(row.get('Doc Type', '')).strip()
                remarks = str(row.get('Remarks', '')).strip()
                status = str(row.get('Status', '')).strip() or 'for signature'
                emp_id = str(row.get('Employee ID', '')).strip()
                vals = [dept, school, emp, code, doc_type, remarks, status]
                col_names = ['department', 'school_office', 'employee', 'code', 'doc_type', 'remarks', 'status']
                if emp_id and emp_id != 'nan':
                    col_names.append('employee_id')
                    vals.append(int(float(emp_id)) if emp_id.replace('.','',1).replace('-','',1).isdigit() else emp_id)
                
                for k, v in row.items():
                    if 'Receiving Office' in k:
                        idx = k.split()[-1]
                        col_names.append(f"receiving_office_{idx}")
                        vals.append(str(v) if v else "")
                    elif 'Receiver Name' in k:
                        idx = k.split()[-1]
                        col_names.append(f"receiver_name_{idx}")
                        vals.append(str(v) if v else "")
                    elif 'Timestamp' in k:
                        idx = k.split()[-1]
                        col_names.append(f"timestamp_{idx}")
                        vals.append(str(v) if v else "")
                        
                placeholders = ['?'] * len(vals)
                update_set = ', '.join([f"{c}=excluded.{c}" for c in col_names if c not in ('code', 'employee', 'id')])
                sql = f"INSERT INTO routing_records ({', '.join(col_names)}) VALUES ({', '.join(placeholders)}) ON CONFLICT(code, employee) DO UPDATE SET {update_set}"
                cursor.execute(sql, vals)
                count += 1
            print(f"GSheets pull: upserted {count} routing records.")
        except gspread.exceptions.WorksheetNotFound:
            print("GSheets pull: 'Routing Records' worksheet not found — skipping.")
            
        # ── 2. Pull Users (upsert — preserves local-only rows) ─────────────────
        try:
            ws_users = sh.worksheet("Users")
            records = ws_users.get_all_records()
            # Preserve the admin account's password hash from GSheets
            for row in records:
                username = str(row.get('Username', '')).strip().lower()
                if not username: continue
                pw_hash = str(row.get('Password Hash', '')).strip()
                role = str(row.get('Role', '')).strip()
                status = str(row.get('Status', '')).strip()
                school_office = str(row.get('School/Office', '')).strip()
                email = str(row.get('Email', '')).strip()
                requires_pw_change = int(row.get('Requires Password Change', 0))
                supervised_schools = str(row.get('Supervised Schools', '')).strip()
                employee_name = str(row.get('Employee Name', '')).strip()
                cursor.execute(
                    "INSERT OR REPLACE INTO users "
                    "(username, password_hash, role, status, school_office, email, requires_password_change, supervised_schools, employee_name) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (username, pw_hash, role, status, school_office, email, requires_pw_change, supervised_schools, employee_name)
                )
            # Ensure admin account exists
            if not cursor.execute("SELECT username FROM users WHERE username = 'admin'").fetchone():
                import hashlib as _hashlib, secrets as _secrets
                fallback_hash = _hashlib.sha256(_secrets.token_urlsafe(12).encode()).hexdigest()
                cursor.execute(
                    "INSERT INTO users (username, password_hash, role, status, school_office, email, requires_password_change, supervised_schools, employee_name) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    ('admin', fallback_hash, 'Admin', 'Approved', '', 'admin@deped.gov.ph', 0, '', '')
                )
            print(f"GSheets pull: users upserted.")
        except gspread.exceptions.WorksheetNotFound:
            print("GSheets pull: 'Users' worksheet not found — skipping.")

        # ── 3. Pull Code Lookup (upsert — preserves local-only rows) ─────────
        try:
            ws_code = sh.worksheet("Code Lookup")
            records = ws_code.get_all_records()
            for row in records:
                code = str(row.get('Code', '')).strip()
                if not code: continue
                dept = str(row.get('Department', '')).strip()
                school = str(row.get('School/Office', '')).strip()
                employees = str(row.get('Employees', '')).strip()
                doc_type = str(row.get('Doc Type', '')).strip()
                gen = str(row.get('Generated', '')).strip()
                cursor.execute(
                    "INSERT OR REPLACE INTO code_lookup "
                    "(code, department, school_office, employees, doc_type, generated_at) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (code, dept, school, employees, doc_type, gen)
                )
            print(f"GSheets pull: code lookups restored.")
        except gspread.exceptions.WorksheetNotFound:
            print("GSheets pull: 'Code Lookup' worksheet not found — skipping.")

        # ── 4. Pull Master Data (from 3 raw sheets) ──────────────────────────
        # Pulls "Master Data Sheet1", "Master Data Sheet2", "Master Data Sheet3"
        # and runs the same transformation logic as init_db() to rebuild master_data.
        _pull_master_data_from_gsheets(sh, cursor)
        
        conn.commit()
        conn.close()
        return count, []
    except Exception as e:
        print(f"pull_all_from_gsheets error: {e}")
        return 0, [str(e)]
    finally:
        _release_sync_lock(lock_fd)


def pull_reference_from_gsheets():
    """Lightweight sync: pull code_lookup and master_data from GSheets using
    upsert (not wipe+reinsert). Used by the scanner tab on load so that
    reference edits made directly in GSheets are available before scanning.
    Returns (success_bool, error_message_or_None)."""
    enabled = get_setting('gsheets_enabled') == 'True'
    if not enabled:
        return False, "GSheets sync disabled."
    sheet_id = get_setting('gsheets_id')
    creds = get_decrypted_setting('gsheets_credentials')
    if not sheet_id or not creds:
        return False, "Missing GSheets credentials."

    try:
        import gspread
        from google.oauth2.service_account import Credentials
        import json

        scopes = ["https://www.googleapis.com/auth/spreadsheets"]
        creds_dict = json.loads(creds)
        credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
        client = gspread.authorize(credentials)
        sh = client.open_by_key(sheet_id)
        conn = get_db_connection()
        cursor = conn.cursor()

        # ── 1. Pull Code Lookup (upsert) ──
        try:
            ws = sh.worksheet("Code Lookup")
            records = ws.get_all_records()
            for row in records:
                code = str(row.get('Code', '')).strip().upper()
                if not code:
                    continue
                dept = str(row.get('Department', '')).strip()
                school = str(row.get('School/Office', '')).strip()
                employees = str(row.get('Employees', '')).strip()
                doc_type = str(row.get('Doc Type', '')).strip()
                gen = str(row.get('Generated', '')).strip()
                cursor.execute(
                    "INSERT OR REPLACE INTO code_lookup "
                    "(code, department, school_office, employees, doc_type, generated_at) "
                    "VALUES (?, ?, ?, ?, ?, ?)",
                    (code, dept, school, employees, doc_type, gen)
                )
        except gspread.exceptions.WorksheetNotFound:
            pass

        # ── 2. Pull Master Data ──
        try:
            _pull_master_data_from_gsheets(sh, cursor)
        except Exception:
            pass

        conn.commit()
        conn.close()
        return True, None
    except Exception as e:
        return False, str(e)


def push_code_to_gsheets(sheet_id, service_account_json, code_dict):
    """Upsert a code_lookup entry into the 'Code Lookup' GSheets worksheet."""
    import gspread
    client = _get_gspread_client(service_account_json)
    sh = client.open_by_key(sheet_id)
    
    try:
        ws = sh.worksheet("Code Lookup")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="Code Lookup", rows=1000, cols=10)
        headers = ["Code", "Department", "School/Office", "Employees", "Doc Type", "Generated"]
        ws.insert_row(headers, 1)
        ws.format("A1:F1", {"backgroundColor": {"red": 0.05, "green": 0.25, "blue": 0.45}, "textFormat": {"foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}, "bold": True}})
        
    code_to_match = str(code_dict.get('code', '')).strip().upper()
    all_rows = ws.get_all_values()
    
    match_row_num = None
    for idx, row in enumerate(all_rows[1:], start=2):
        if len(row) > 0 and str(row[0]).strip().upper() == code_to_match:
            match_row_num = idx
            break
            
    row_data = [
        code_dict.get('code', ''),
        code_dict.get('department', ''),
        code_dict.get('school_office', ''),
        code_dict.get('employees', ''),
        code_dict.get('doc_type', ''),
        code_dict.get('generated_at', '')
    ]
    
    if match_row_num:
        ws.update(range_name=f"A{match_row_num}:F{match_row_num}", values=[row_data])
    else:
        ws.append_row(row_data)


def push_schools_to_gsheets(sheet_id, service_account_json):
    """Full-snapshot push: clear + rewrite 'Schools' worksheet."""
    import gspread
    client = _get_gspread_client(service_account_json)
    sh = client.open_by_key(sheet_id)
    try:
        ws = sh.worksheet("Schools")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="Schools", rows=1000, cols=5)

    conn = get_db_connection()
    rows = conn.execute("SELECT id, name, department FROM schools ORDER BY id").fetchall()
    conn.close()

    headers = ["id", "name", "department"]
    data = [headers]
    for r in rows:
        data.append([r['id'], r['name'], r['department']])

    ws.clear()
    if data:
        end_col = _col_to_letter(len(headers))
        ws.update(range_name=f"A1:{end_col}{len(data)}", values=data)
        ws.format(f"A1:{end_col}1", {
            "backgroundColor": {"red": 0.05, "green": 0.25, "blue": 0.45},
            "textFormat": {"foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}, "bold": True}
        })
        ws.freeze(rows=1)


def push_employees_to_gsheets(sheet_id, service_account_json):
    """Full-snapshot push: clear + rewrite 'Employees' worksheet."""
    import gspread
    client = _get_gspread_client(service_account_json)
    sh = client.open_by_key(sheet_id)
    try:
        ws = sh.worksheet("Employees")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="Employees", rows=20000, cols=5)

    conn = get_db_connection()
    rows = conn.execute("SELECT e.id, e.name, e.school_id, s.name AS school_name FROM employees e JOIN schools s ON s.id = e.school_id ORDER BY e.id").fetchall()
    conn.close()

    headers = ["id", "name", "school_id", "school_name"]
    data = [headers]
    for r in rows:
        data.append([r['id'], r['name'], r['school_id'], r['school_name']])

    ws.clear()
    if data:
        end_col = _col_to_letter(len(headers))
        ws.update(range_name=f"A1:{end_col}{len(data)}", values=data)
        ws.format(f"A1:{end_col}1", {
            "backgroundColor": {"red": 0.05, "green": 0.25, "blue": 0.45},
            "textFormat": {"foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}, "bold": True}
        })
        ws.freeze(rows=1)


def push_code_employees_to_gsheets(sheet_id, service_account_json):
    """Append-only push: only new/modified rows since last push.
    Rows with active=FALSE are toggled in-place (set column C to FALSE)."""
    import gspread
    import datetime
    client = _get_gspread_client(service_account_json)
    sh = client.open_by_key(sheet_id)
    try:
        ws = sh.worksheet("Code Employees")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="Code Employees", rows=10000, cols=5)
        headers = ["code", "employee_id", "active", "updated_at"]
        ws.insert_row(headers, 1)
        ws.format("A1:D1", {
            "backgroundColor": {"red": 0.05, "green": 0.25, "blue": 0.45},
            "textFormat": {"foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}, "bold": True}
        })

    # Get all existing code-employee combos from sheet to avoid re-appending active rows
    existing = set()
    all_rows = ws.get_all_values()
    for row in all_rows[1:]:
        if len(row) >= 2:
            existing.add((str(row[0]).strip(), str(row[1]).strip()))

    now = datetime.datetime.utcnow().isoformat()
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT ce.code, ce.employee_id, 1 AS active, ? AS updated_at
        FROM code_employees ce
        ORDER BY ce.code, ce.employee_id
    """, (now,)).fetchall()
    conn.close()

    append_data = []
    for r in rows:
        key = (str(r['code']).strip(), str(r['employee_id']).strip())
        if key not in existing:
            append_data.append([r['code'], r['employee_id'], 'TRUE', r['updated_at']])

    if append_data:
        ws.append_rows(append_data)


def push_master_db_to_gsheets(sheet_id, service_account_json):
    """Push all 3 raw sheets from master_db.xlsx to Google Sheets as separate worksheets.
    
    Sheet1 → 'Master Data Sheet1'  (Department, School/Office, Employee_Name)
    Sheet2 → 'Master Data Sheet2'  (FIRSTNAME, MIDDLE INITIAL, LASTNAME, SUFFIX, SCHOOL/ OFFICE, ...)
    Sheet3 → 'Master Data Sheet3'  (Senior High Schools list)
    """
    import gspread
    if not os.path.exists(MASTER_DB):
        print("GSheets push: master_db.xlsx not found — skipping master sheet push.")
        return

    xl = pd.ExcelFile(MASTER_DB)
    client = _get_gspread_client(service_account_json)
    sh = client.open_by_key(sheet_id)

    # Helper to style a worksheet header row
    def _style_header(ws, num_cols):
        end_col = _col_to_letter(num_cols)
        ws.format(f"A1:{end_col}1", {
            "backgroundColor": {"red": 0.05, "green": 0.25, "blue": 0.45},
            "horizontalAlignment": "CENTER",
            "textFormat": {
                "foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0},
                "bold": True, "fontSize": 11
            }
        })
        ws.freeze(rows=1)

    # Helper to push a DataFrame to a named worksheet
    def _push_sheet(worksheet_name, df, num_rows_estimate=None):
        try:
            ws = sh.worksheet(worksheet_name)
        except gspread.exceptions.WorksheetNotFound:
            rows_needed = max(1000, (num_rows_estimate or len(df)) + 100)
            ws = sh.add_worksheet(title=worksheet_name, rows=rows_needed, cols=max(len(df.columns), 5))

        headers = list(df.columns.astype(str))
        rows_to_write = [headers]
        for _, row in df.iterrows():
            rows_to_write.append([str(v) if pd.notna(v) else '' for v in row])

        ws.clear()
        if rows_to_write:
            end_col = _col_to_letter(len(headers))
            ws.update(range_name=f"A1:{end_col}{len(rows_to_write)}", values=rows_to_write)
        _style_header(ws, len(headers))
        print(f"GSheets push: '{worksheet_name}' — {len(df)} rows pushed.")

    # 1. Push Sheet1
    if 'Sheet1' in xl.sheet_names:
        df1 = pd.read_excel(xl, 'Sheet1')
        _push_sheet('Master Data Sheet1', df1)

    # 2. Push Sheet2
    if 'Sheet2' in xl.sheet_names:
        df2 = pd.read_excel(xl, 'Sheet2')
        _push_sheet('Master Data Sheet2', df2)

    # 3. Push Sheet3
    if 'Sheet3' in xl.sheet_names:
        df3 = pd.read_excel(xl, 'Sheet3')
        _push_sheet('Master Data Sheet3', df3)


def push_user_to_gsheets(sheet_id, service_account_json, user_dict):
    """Upsert a user record into the 'Users' GSheets worksheet.
    Stores all user fields including requires_password_change so it survives restarts."""
    import gspread
    client = _get_gspread_client(service_account_json)
    sh = client.open_by_key(sheet_id)
    
    try:
        ws = sh.worksheet("Users")
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title="Users", rows=1000, cols=10)
        headers = ["Username", "Password Hash", "Role", "Status", "School/Office", "Email", "Requires Password Change", "Supervised Schools", "Employee Name"]
        ws.insert_row(headers, 1)
        ws.format("A1:I1", {"backgroundColor": {"red": 0.05, "green": 0.25, "blue": 0.45}, "textFormat": {"foregroundColor": {"red": 1.0, "green": 1.0, "blue": 1.0}, "bold": True}})
        
    username = str(user_dict.get('username', '')).strip().lower()
    all_rows = ws.get_all_values()

    # Ensure the header row has the requires_password_change column
    existing_headers = all_rows[0] if all_rows else []
    if existing_headers and "Requires Password Change" not in existing_headers:
        col_idx = len(existing_headers) + 1
        ws.update_cell(1, col_idx, "Requires Password Change")
        existing_headers.append("Requires Password Change")

    # Ensure the header row has the Employee Name column
    if existing_headers and "Employee Name" not in existing_headers:
        col_idx = len(existing_headers) + 1
        ws.update_cell(1, col_idx, "Employee Name")
        existing_headers.append("Employee Name")
    
    match_row_num = None
    for idx, row in enumerate(all_rows[1:], start=2):
        if len(row) > 0 and str(row[0]).strip().lower() == username:
            match_row_num = idx
            break
            
    row_data = [
        user_dict.get('username', ''),
        '',  # password_hash intentionally blank — never pushed to sheets
        user_dict.get('role', ''),
        user_dict.get('status', ''),
        user_dict.get('school_office', ''),
        user_dict.get('email', ''),
        user_dict.get('requires_password_change', 0),
        user_dict.get('supervised_schools', ''),
        user_dict.get('employee_name', ''),
    ]
    
    if match_row_num:
        ws.update(range_name=f"A{match_row_num}:I{match_row_num}", values=[row_data])
    else:
        ws.append_row(row_data)


# --- SCHOOL CODE MERGE HELPERS ---

def find_duplicate_codes(doc_type):
    """Find schools with multiple generated codes for a given doc_type.
    Returns list of {school_office, codes: [{code, dept, employees, routing_count, generated_at}]}."""
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT code, department, school_office, employees, generated_at, doc_type FROM code_lookup WHERE doc_type = ? ORDER BY school_office, generated_at DESC",
        (doc_type,)
    ).fetchall()
    conn.close()

    school_groups = {}
    for r in rows:
        d = dict(r)
        school = d['school_office']
        if school not in school_groups:
            school_groups[school] = []
        school_groups[school].append(d)

    result = []
    conn2 = get_db_connection()
    for school, codes in school_groups.items():
        if len(codes) < 2:
            continue
        # Enrich with routing_count and determine survivor (code with most routing records)
        enriched = []
        for c in codes:
            cnt = conn2.execute("SELECT COUNT(*) FROM routing_records WHERE code = ?", (c['code'],)).fetchone()[0]
            enriched.append({**c, 'routing_count': cnt})
        enriched.sort(key=lambda x: x['routing_count'], reverse=True)
        result.append({
            'school_office': school,
            'department': enriched[0]['department'],
            'codes': enriched,
            'survivor_code': enriched[0]['code']
        })
    conn2.close()
    return result


def merge_codes(doc_type, merges):
    """Execute merge for given merges list.
    merges: [{school_office, survivor_code, delete_codes}]
    Returns {merged_count, reassigned_records, deleted_codes}."""
    conn = get_db_connection()
    total_reassigned = 0
    deleted_codes_list = []
    merged_count = len(merges)

    for m in merges:
        survivor = m['survivor_code']
        to_delete = m.get('delete_codes', [])

        # Reassign routing records from deleted codes to survivor
        for old_code in to_delete:
            # Count records to be reassigned
            cnt = conn.execute(
                "SELECT COUNT(*) FROM routing_records WHERE code = ? AND code != ?",
                (old_code, survivor)
            ).fetchone()[0]
            if cnt:
                conn.execute(
                    "UPDATE routing_records SET code = ? WHERE code = ? AND code != ?",
                    (survivor, old_code, survivor)
                )
                total_reassigned += cnt
            # Delete the old code from code_lookup
            conn.execute("DELETE FROM code_lookup WHERE code = ?", (old_code,))
            deleted_codes_list.append(old_code)

    conn.commit()
    conn.close()

    # Queue sync for all affected routing records
    queue_sync_batch(['all'])

    return {
        'merged_count': merged_count,
        'reassigned_records': total_reassigned,
        'deleted_codes': deleted_codes_list
    }


# ─────────────────────────────────────────────
#  PENDING PROFILE CHANGES (admin approval)
# ─────────────────────────────────────────────

def create_pending_change(username, changes_dict):
    """Insert a pending profile change request. Returns the new row id."""
    conn = get_db_connection()
    now = datetime.now().isoformat()
    cursor = conn.cursor()
    cursor.execute(
        "INSERT INTO pending_profile_changes (username, changes_json, requested_at, status) VALUES (?, ?, ?, 'pending')",
        (username, json.dumps(changes_dict), now)
    )
    new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return new_id


def get_pending_changes():
    """Return all pending_profile_changes rows ordered by newest first."""
    conn = get_db_connection()
    rows = conn.execute(
        "SELECT * FROM pending_profile_changes WHERE status = 'pending' ORDER BY requested_at DESC"
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_pending_change(change_id):
    """Return a single pending change by id."""
    conn = get_db_connection()
    row = conn.execute(
        "SELECT * FROM pending_profile_changes WHERE id = ?", (change_id,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def apply_pending_change(change_id):
    """Apply the changes to the users table and mark as approved.
    Returns {'status': 'ok', 'changes_applied': [...]} or {'status': 'error', 'message': ...}
    """
    conn = get_db_connection()
    row = conn.execute(
        "SELECT * FROM pending_profile_changes WHERE id = ? AND status = 'pending'",
        (change_id,)
    ).fetchone()
    if not row:
        conn.close()
        return {'status': 'error', 'message': 'Pending change not found or already processed.'}

    changes = json.loads(row['changes_json'])
    username = row['username']
    applied = []

    # Build update SET clause
    update_fields = {}
    for field in ['employee_name', 'email', 'supervised_schools', 'username']:
        if field in changes:
            new_val = changes[field]
            # Skip if null/empty
            if new_val is None or (isinstance(new_val, str) and not new_val.strip()):
                continue
            update_fields[field] = new_val.strip() if isinstance(new_val, str) else new_val
            applied.append(field)

    # Also handle password_hash change (applied immediately by request_profile_change endpoint)
    # No need to apply here since it was already done.

    if update_fields:
        set_clause = ', '.join([f"{k} = ?" for k in update_fields.keys()])
        values = list(update_fields.values())
        # Update the users table
        # If username is being changed, the current username is the old one
        current_username = changes.get('_current_username', username)
        conn.execute(f"UPDATE users SET {set_clause} WHERE username = ?", values + [current_username])
        applied.append('users_table_updated')

    conn.execute(
        "UPDATE pending_profile_changes SET status = 'approved' WHERE id = ?",
        (change_id,)
    )
    conn.commit()
    conn.close()
    return {'status': 'ok', 'changes_applied': applied}


def reject_pending_change(change_id):
    """Mark a pending change as rejected."""
    conn = get_db_connection()
    conn.execute(
        "UPDATE pending_profile_changes SET status = 'rejected' WHERE id = ?",
        (change_id,)
    )
    conn.commit()
    conn.close()
    return {'status': 'ok'}
